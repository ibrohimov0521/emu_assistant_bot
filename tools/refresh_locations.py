"""EMU API dan hudud ma'lumotlarini olib `location_data.py` faylini yangilaydi.

Ishlatish:

    python tools/refresh_locations.py

Skript uchta manbani birlashtiradi:

* `https://apiv1.emu.uz/api/v1/regions`  - 14 viloyat
* `https://apiv1.emu.uz/api/v1/cities`   - 198 shahar/tuman (`extra_name` = server formati)
* `https://apiv1.emu.uz/api/v1/branches` - har bir ofis qaysi tumanda joylashgani
* `templates/branch_codes.xlsx`          - ofis nomi -> ichki filial kodi

Natija: `location_data.py` (qo'lda tahrirlanmaydi). Qo'lda kiritiladigan taxalluslar
(`Denau`, `Zafar`, `Toʻytepa` va boshqalar) `locations.py` ichida saqlanadi.
"""

from __future__ import annotations

import json
import re
import sys
import urllib.request
from datetime import date
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

# Diqqat: bu skript `locations.py` ni import qilmaydi, chunki u shu skript yaratadigan
# `location_data.py` ga tayanadi. Shu sabab kalit yasash funksiyasi shu yerda takrorlanadi.
CYRILLIC_TO_LATIN = str.maketrans(
    {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo", "ж": "j", "з": "z",
        "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
        "с": "s", "т": "t", "у": "u", "ф": "f", "х": "x", "ц": "s", "ч": "ch", "ш": "sh", "щ": "sh",
        "ъ": "", "ы": "i", "ь": "", "э": "e", "ю": "yu", "я": "ya", "қ": "q", "ғ": "g", "ҳ": "h", "ў": "o",
    }
)


def normalize_location_key(value: str) -> str:
    text = str(value or "").casefold().translate(CYRILLIC_TO_LATIN)
    for mark in ("ʼ", "ʻ", "‘", "’", "'", "`", "´"):
        text = text.replace(mark, "")
    return re.sub(r"[^a-z0-9]+", "", text)


API_BASE_URL = "https://apiv1.emu.uz"
BRANCH_CODES_PATH = BASE_DIR / "templates" / "branch_codes.xlsx"
OUTPUT_PATH = BASE_DIR / "location_data.py"

# Viloyat markazi bo'lgan shahar id lari.
REGION_CENTER_CITY_ID = {
    1: 1,      # Andijon
    2: 3,      # Buxoro
    3: 11,     # Farg'ona
    4: 4,      # Jizzax
    5: 5,      # Qarshi
    6: 10,     # Urganch
    7: 7,      # Namangan
    8: 6,      # Navoiy
    9: 8,      # Nukus
    10: 2,     # Samarqand
    11: 9,     # Termiz
    12: 12,    # Guliston
    13: 198,   # Toshkent shahri
    14: 182,   # Nurafshon (O'rtachirchiq tumani)
}

# EMU API `extra_name` bilan import dasturidagi справочник orasidagi imlo farqlari.
# Chapda API yozuvi, o'ngda bot ilgari ishlatgan (va server qabul qilgan) yozuv.
SERVER_NAME_OVERRIDES = {
    "Сергелий": "Сергели",
    "Янгиюль": "Янгийоль",
    "Казанкеткен": "Казакеткен",
    "Баландчакыр": "Балангачкыр",
    "Верхне-чирчикский": "Верхне-Чирчикский",
}

# branch_codes.xlsx dagi nomlar API ofis ro'yxatida uchramaydi - qo'lda bog'lanadi.
OFFICE_CITY_OVERRIDES = {
    "Qorasuv": "Xonobod shahri",
    "Samarqand Qorasuv": "Samarqand shahri",
    "Ippodrom": "Chilonzor tumani",
    "Tinchlik": "Olmazor tumani",
    "Toshkent Courier": "Toshkent",
}


def api_get(path: str) -> list[dict]:
    request = urllib.request.Request(
        f"{API_BASE_URL}{path}",
        headers={"Accept": "application/json", "User-Agent": "emu_assistant_bot/1.0"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def localized(item: dict, field: str, language: str) -> str:
    values = item.get(field) or {}
    return str(values.get(language) or "").strip()


def load_branch_names() -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(BRANCH_CODES_PATH, data_only=True)
    try:
        sheet = workbook.active
        names = []
        for row in range(2, sheet.max_row + 1):
            code = sheet.cell(row, 1).value
            name = sheet.cell(row, 3).value
            if code and name:
                names.append(str(name).strip())
        return names
    finally:
        workbook.close()


def python_literal(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def main() -> int:
    regions = api_get("/api/v1/regions")
    cities = api_get("/api/v1/cities")

    branch_city_by_name: dict[str, str] = {}
    for region in regions:
        for branch in api_get(f"/api/v1/branches?region_id={region['id']}"):
            key = normalize_location_key(str(branch.get("name") or ""))
            city_name = str(branch.get("city_name") or "").strip()
            if key and city_name:
                branch_city_by_name.setdefault(key, city_name)

    city_by_uz_key = {normalize_location_key(localized(city, "i18n_name", "UZ")): city for city in cities}

    offices_by_city_id: dict[int, list[str]] = {}
    unresolved: list[str] = []
    for name in load_branch_names():
        city_name = OFFICE_CITY_OVERRIDES.get(name) or branch_city_by_name.get(normalize_location_key(name))
        city = city_by_uz_key.get(normalize_location_key(city_name or ""))
        if city is None:
            unresolved.append(name)
            continue
        offices_by_city_id.setdefault(int(city["id"]), []).append(name)

    if unresolved:
        print("DIQQAT: quyidagi ofislar tumanga bog'lanmadi:", ", ".join(unresolved))

    lines = [
        '"""EMU API asosida avtomatik yaratilgan hudud ma\'lumotlari.',
        "",
        "Bu faylni qo'lda tahrirlamang - `python tools/refresh_locations.py` bilan yangilanadi.",
        "Manba: https://apiv1.emu.uz (regions, cities, branches) + templates/branch_codes.xlsx",
        '"""',
        "",
        f"GENERATED_AT = {python_literal(date.today().isoformat())}",
        "",
        "# (region_id, uz_name, ru_name, center_city_id)",
        "REGIONS = (",
    ]
    for region in sorted(regions, key=lambda item: int(item["id"])):
        region_id = int(region["id"])
        lines.append(
            "    ({}, {}, {}, {}),".format(
                region_id,
                python_literal(str(region.get("name") or "").strip()),
                python_literal(localized(region, "i18n_name", "RU")),
                REGION_CENTER_CITY_ID[region_id],
            )
        )
    lines += [
        ")",
        "",
        "# (city_id, region_id, uz_name, ru_name, server_name, offices)",
        "CITIES = (",
    ]
    for city in sorted(cities, key=lambda item: (int(item["region_id"]), int(item["id"]))):
        city_id = int(city["id"])
        extra_name = str(city.get("extra_name") or "").strip()
        server_name = SERVER_NAME_OVERRIDES.get(extra_name, extra_name)
        offices = offices_by_city_id.get(city_id, [])
        offices_literal = "(" + "".join(f"{python_literal(name)}, " for name in offices) + ")" if offices else "()"
        lines.append(
            "    ({}, {}, {}, {}, {}, {}),".format(
                city_id,
                int(city["region_id"]),
                python_literal(localized(city, "i18n_name", "UZ")),
                python_literal(re.sub(r"\s+", " ", localized(city, "i18n_name", "RU"))),
                python_literal(server_name),
                offices_literal,
            )
        )
    lines.append(")")
    lines.append("")

    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"{OUTPUT_PATH.name}: {len(regions)} viloyat, {len(cities)} shahar/tuman yozildi")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
