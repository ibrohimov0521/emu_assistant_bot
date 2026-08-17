"""Hudud aniqlash va filial kodi testlari.

Ishlatish (aiogram/openai kerak emas):

    python -m unittest discover -s tests -v
"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

import locations as L  # noqa: E402
from location_data import CITIES  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

BRANCH_CODES_PATH = BASE_DIR / "templates" / "branch_codes.xlsx"


def load_branch_codes() -> dict[str, str]:
    """branch_codes.xlsx: normallashtirilgan ofis nomi -> ichki kod."""

    workbook = load_workbook(BRANCH_CODES_PATH, data_only=True)
    try:
        sheet = workbook.active
        codes = {}
        for row in range(2, sheet.max_row + 1):
            code = sheet.cell(row, 1).value
            name = sheet.cell(row, 3).value
            if code and name:
                codes.setdefault(L.normalize_location_key(str(name)), str(code).strip())
        return codes
    finally:
        workbook.close()


BRANCH_CODES = load_branch_codes()


class DataIntegrityTests(unittest.TestCase):
    def test_tables_are_valid(self):
        self.assertEqual([], L.validate())

    def test_all_regions_and_cities_present(self):
        self.assertEqual(14, len(L.REGION_BY_ID))
        self.assertEqual(198, len(L.PLACES))
        self.assertEqual(len(L.PLACES), len(L.SERVER_LOCATIONS))

    def test_every_region_has_at_least_one_office_code(self):
        for region in L.REGION_BY_ID.values():
            center = L.PLACE_BY_ID[region.center_city_id]
            names = L.region_offices_for_server(center.server)
            codes = [BRANCH_CODES.get(L.normalize_location_key(name)) for name in names]
            self.assertTrue(
                any(code for code in codes),
                f"{region.uz}: viloyat bo'yicha birorta filial kodi topilmadi",
            )

    def test_extra_names_point_to_known_cities(self):
        known = {L.normalize_location_key(uz) for _, _, uz, _, _, _ in CITIES}
        for table in (L.EXTRA_NAMES, L.WEAK_NAMES):
            for uz_name in table:
                self.assertIn(L.normalize_location_key(uz_name), known, uz_name)

    def test_offices_exist_in_branch_codes_file(self):
        """`location_data.py` dagi ofislar branch_codes.xlsx da bo'lishi kerak."""

        missing = []
        for place in L.PLACES:
            for name in place.offices:
                if L.normalize_location_key(name) not in BRANCH_CODES:
                    missing.append(f"{place.uz}: {name}")
        self.assertEqual([], missing)


class NameCoverageTests(unittest.TestCase):
    """EMU ro'yxatidagi har bir nom o'z hududiga qaytishi kerak."""

    def test_uzbek_names(self):
        for place in L.PLACES:
            with self.subTest(place=place.uz):
                self.assertEqual(place.server, L.resolve_location(place.uz).server)

    def test_russian_names(self):
        for place in L.PLACES:
            with self.subTest(place=place.ru):
                self.assertEqual(place.server, L.resolve_location(place.ru).server)

    def test_server_names(self):
        for place in L.PLACES:
            with self.subTest(place=place.server):
                self.assertEqual(place.server, L.resolve_location(place.server).server)

    def test_short_uzbek_names_with_region_hint(self):
        """'Denov' (tuman so'zi yozilmagan) + viloyat nomi ham topilishi kerak."""

        for place in L.PLACES:
            base = L._uz_base_name(place.uz)
            if not base:
                continue
            text = f"{place.region.uz}, {base}"
            with self.subTest(text=text):
                self.assertTrue(L.resolve_location(text).server, text)

    def test_every_district_gets_a_branch_code(self):
        """Har bir tuman uchun filial kodi chiqishi kerak (o'zi yoki viloyat markazi)."""

        for place in L.PLACES:
            names = L.offices_for_server(place.server) or L.region_offices_for_server(place.server)
            codes = [BRANCH_CODES.get(L.normalize_location_key(name)) for name in names]
            with self.subTest(place=place.server):
                self.assertTrue(any(code for code in codes), f"{place.server}: kod topilmadi")


class DenovTests(unittest.TestCase):
    """Foydalanuvchi aytgan holat: Denov / денов / ДЕНОУ -> Денау."""

    def test_all_denov_spellings(self):
        for text in (
            "denov",
            "Denov",
            "денов",
            "ДЕНОВ",
            "ДЕНОУ",
            "Деноу",
            "Denau",
            "Денау",
            "Denov tumani",
            "денов тумани",
            "Денауский район",
            "Surxondaryo viloyati, Denov tumani",
            "Сурхандарьинская область, Денов",
            "surxondaryo denov tumani beshkapa",
        ):
            with self.subTest(text=text):
                self.assertEqual("Денау", L.resolve_location(text).server)

    def test_denov_office_code(self):
        offices = L.offices_for_server("Денау")
        self.assertIn("Denov O'zbegim", offices)
        self.assertEqual("23", BRANCH_CODES[L.normalize_location_key("Denov O'zbegim")])


class ResolutionTests(unittest.TestCase):
    def check(self, text: str, expected: str):
        with self.subTest(text=text):
            self.assertEqual(expected, L.resolve_location(text).server)

    def test_mixed_scripts_and_transliterations(self):
        cases = {
            "Toshkent": "Ташкент",
            "Tashkent": "Ташкент",
            "Ташкент шахри": "Ташкент",
            "samarkand": "Самарканд",
            "Samarqand shahri": "Самарканд",
            "Bukhara": "Бухара",
            "buxoro shahri": "Бухара",
            "Khiva": "Хива",
            "Xiva tumani": "Хива",
            "Fergana": "Фергана",
            "Farg'ona shahri": "Фергана",
            "Kokand": "Коканд",
            "Qo'qon": "Коканд",
            "Jizzakh": "Джизак",
            "Djizak": "Джизак",
            "Termez": "Термез",
            "Termiz shahri": "Термез",
            "Urgench": "Ургенч",
            "Urganch shahri": "Ургенч",
            "Karshi": "Карши",
            "Qarshi shahri": "Карши",
            "Navoi": "Навои",
            "Нукус": "Нукус",
            "Andijan": "Андижан",
            "Gulistan": "Гулистан",
        }
        for text, expected in cases.items():
            self.check(text, expected)

    def test_district_centers_and_old_names(self):
        cases = {
            "Oltinsoy tumani": "Карлук",
            "Qorluq": "Карлук",
            "Muzrabot": "Халкабад",
            "Xalqobod": "Халкабад",
            "Forish tumani": "Янгикишлак",
            "Yangiqishloq": "Янгикишлак",
            "Mirzacho'l tumani": "Гагарин",
            "Furqat tumani": "Навбахор",
            "Buvayda": "Бувайда",
            "Ibrat": "Бувайда",
            "To'ytepa": "Нурафшон",
            "O'rtachirchiq tumani": "Нурафшон",
            "Zafar": "Бекабад",
            "Bekobod": "Бекабад",
            "Ellikqala": "Элликкала",
            "Bo'ston": "Элликкала",
            "Mang'it": "Амударья",
            "Amudaryo tumani": "Амударья",
            "Payshanba": "Каттакурган",
            "Ziyovuddin": "Зиадин",
            "Yaypan": "Яйпан",
            "O'zbekiston tumani": "Яйпан",
            "Keles": "Келес",
            "Toshkent tumani": "Келес",
            "Pitnak": "Тупраккала",
            "Gazalkent": "Газалкент",
            "Bo'stonliq tumani": "Газалкент",
        }
        for text, expected in cases.items():
            self.check(text, expected)

    def test_full_addresses(self):
        cases = {
            "Surxondaryo viloyati, Denov tumani, Sharof Rashidov ko'chasi 12": "Денау",
            "Toshkent shahri, Yunusobod tumani, 12-kvartal, 5-uy": "Юнусабад",
            "Toshkent viloyati, Bekobod tumani, Navoiy ko'chasi 4": "Бекабад",
            "Farg'ona viloyati, Quva tumani, Ahmad Yassaviy ko'chasi 44": "Кува",
            "Андижанская область, Асакинский район, ул. Умид 90": "Асака",
            "Xorazm viloyati Xiva shahri Amir Temur ko'chasi 13": "Хива",
            "Namangan viloyati, Chust tumani, Ipak yo'li ko'chasi 7": "Чуст",
        }
        for text, expected in cases.items():
            self.check(text, expected)

    def test_region_only_falls_back_to_center(self):
        match = L.resolve_location("Surxondaryo viloyati")
        self.assertEqual("Термез", match.server)
        self.assertTrue(match.note)

    def test_ambiguous_names_use_region_hint(self):
        self.assertEqual("Акалтын", L.resolve_location("Oqoltin tumani").server)
        self.assertEqual("Улугнор", L.resolve_location("Andijon viloyati, Oqoltin").server)
        self.assertEqual("Дехканабадский район", L.resolve_location("Dehqonobod tumani").server)
        self.assertEqual("Дехканабад", L.resolve_location("Sirdaryo viloyati, Dehqonobod").server)
        self.assertEqual("Янгибазар", L.resolve_location("Yangibozor tumani").server)
        self.assertEqual("Верхне-Чирчикский", L.resolve_location("Toshkent viloyati, Yangibozor").server)

    def test_street_names_are_not_treated_as_districts(self):
        # Faqat ko'cha nomi bo'lsa - hudud taxmin qilinmaydi, izoh qoldiriladi.
        match = L.resolve_location("улица Шарафа Рашидова, 242")
        self.assertTrue(match.note)
        self.assertTrue(match.approximate or not match.server)
        # Shahar nomi bo'lsa - ko'cha nomi emas, shahar tanlanadi.
        self.assertEqual("Термез", L.resolve_location("Termiz shahri, Sharof Rashidov ko'chasi 242").server)
        self.assertEqual("Бухара", L.resolve_location("Buxoro shahri, Samarqand ko'chasi 12").server)

    def test_city_and_district_with_the_same_name(self):
        """'Buxoro shahri' -> Бухара, 'Buxoro tumani' -> Галлаасия."""

        cases = {
            "Buxoro shahri, Samarqand ko'chasi 15": "Бухара",
            "Buxoro tumani": "Галлаасия",
            "Qarshi shahri, Islom Karimov ko'chasi 295": "Карши",
            "Qarshi tumani": "Бешкент",
            "Termiz shahri, Alisher Navoiy 25": "Термез",
            "Termiz tumani": "Учкизил",
            "Namangan shahri, Lola mahallasi": "Наманган",
            "Namangan tumani": "Ташбулак",
            "Guliston shahri, Toshkent yo'li 1": "Гулистан",
            "Guliston tumani": "Дехканабад",
            "Urganch shahri, Yog'du ko'chasi 8": "Ургенч",
            "Urganch tumani": "Караул",
            "Toshkent shahri": "Ташкент",
            "Toshkent tumani": "Келес",
        }
        for text, expected in cases.items():
            self.check(text, expected)

    def test_district_wins_over_its_own_region_center(self):
        """Matnda ham tuman, ham viloyat markazi bo'lsa - tuman aniqroq."""

        self.check("Yunusobod 12-kvartal 5-uy, Toshkent", "Юнусабад")
        self.check("Toshkent shahri, Sergeli tumani, massiv Sergeli-8A", "Сергели")
        self.check("Toshkent, Chilonzor 19-kvartal", "Чиланзар")

    def test_typos_are_matched_approximately(self):
        match = L.resolve_location("Denou tumani")
        self.assertEqual("Денау", match.server)
        self.assertEqual("Самарканд", L.resolve_location("Samarqannd shahri").server)

    def test_unknown_text_returns_note(self):
        match = L.resolve_location("qwerty 12345")
        self.assertEqual("", match.server)
        self.assertTrue(match.note)

    def test_first_confident_text_wins(self):
        match = L.resolve_location("", "Denov tumani", "Ташкент")
        self.assertEqual("Денау", match.server)
        match = L.resolve_location("12-uy", "", "Ташкент")
        self.assertEqual("Ташкент", match.server)


class KeyHelperTests(unittest.TestCase):
    def test_normalize_location_key(self):
        self.assertEqual("qoqon", L.normalize_location_key("Qo‘qon"))
        self.assertEqual("qoqon", L.normalize_location_key("Qo'qon"))
        self.assertEqual("denov", L.normalize_location_key("ДЕНОВ"))
        self.assertEqual("mirzoulugbek", L.normalize_location_key("Мирзо-Улугбек"))

    def test_phonetic_key_merges_spelling_variants(self):
        self.assertEqual(L.phonetic_key("xiva"), L.phonetic_key("khiva"))
        self.assertEqual(L.phonetic_key("qarshi"), L.phonetic_key("karshi"))
        self.assertEqual(L.phonetic_key("jizzax"), L.phonetic_key("jizzakh"))

    def test_location_tokens_drops_markers(self):
        self.assertEqual(["denov"], L.location_tokens("Denov tumani"))
        self.assertEqual(["navoiy"], L.location_tokens("Navoiy ko'chasi"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
