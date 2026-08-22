"""`templates/branch_codes.xlsx` bo'yicha ichki filial kodini tanlash.

Hudud (`Place.server`, masalan "Денау") aniqlangandan keyin ДО ОФИСА jo'natmalari
uchun D ustunga shu tumandagi ofisning ichki kodi yoziladi.

Tanlash tartibi:

1. `location_data.py` da shu tumanga bog'langan ofis(lar) - manzilga eng mos keluvchisi.
2. Nomi tuman nomiga to'g'ri keladigan ofis (fayl yangilangan, lekin `location_data.py`
   hali yangilanmagan holat uchun zaxira).
3. Shu tuman/shaharda ofis bo'lmasa bo'sh kod qaytadi; asosiy dastur bu qatorni
   НА ДОМ qilib, tuman markaziga yozadi.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from difflib import get_close_matches
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from locations import (
    location_tokens,
    location_words,
    name_keys_for_server,
    normalize_location_key,
    offices_for_server,
    region_center_server,
    region_offices_for_server,
)

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
BRANCH_CODES_PATH = BASE_DIR / "templates" / "branch_codes.xlsx"

# Manzilda uchraydigan, filialni ajratishga yordam bermaydigan so'zlar.
ADDRESS_GENERIC_TOKENS = {
    "viloyati",
    "viloyat",
    "shahar",
    "shaxar",
    "tumani",
    "tuman",
    "rayoni",
    "rayon",
    "kochasi",
    "kocha",
    "ulisa",
    "ulitsa",
    "massiv",
    "mahalla",
    "mahallasi",
    "mfy",
    "uy",
    "dom",
}

CENTRAL_BRANCH_TOKENS = {
    "uzbekistan",
    "uzbekistanskaya",
    "uzbekistanski",
    "uzbekiston",
    "mustaqillik",
    "mustakillik",
    "istiklol",
    "navoi",
    "amirtemur",
    "markaz",
    "center",
    "central",
    "sentral",
    "vokzal",
    "station",
}


@dataclass(frozen=True)
class BranchCodeRecord:
    code: str
    parent: str
    name: str
    address: str


_records_cache: list[BranchCodeRecord] | None = None


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def load_branch_code_records() -> list[BranchCodeRecord]:
    global _records_cache
    if _records_cache is not None:
        return _records_cache

    records: list[BranchCodeRecord] = []
    if not BRANCH_CODES_PATH.exists():
        logger.warning("Branch code file not found: %s", BRANCH_CODES_PATH)
        _records_cache = records
        return _records_cache

    workbook = load_workbook(BRANCH_CODES_PATH, data_only=True)
    try:
        sheet = workbook.active
        for row_index in range(2, sheet.max_row + 1):
            code = _text(sheet.cell(row_index, 1).value)
            if not code:
                continue

            if sheet.max_column >= 4:
                parent = _text(sheet.cell(row_index, 2).value)
                name = _text(sheet.cell(row_index, 3).value)
                address = _text(sheet.cell(row_index, 4).value)
            else:
                parent = ""
                name = _text(sheet.cell(row_index, 2).value)
                address = ""

            if name:
                records.append(BranchCodeRecord(code=code, parent=parent, name=name, address=address))
    finally:
        workbook.close()

    _records_cache = records
    return _records_cache


def reset_cache() -> None:
    global _records_cache
    _records_cache = None


def branch_records_by_name() -> dict[str, BranchCodeRecord]:
    return {normalize_location_key(record.name): record for record in load_branch_code_records()}


def branch_record_text(record: BranchCodeRecord) -> str:
    return " ".join([record.parent, record.name, record.address])


def branch_record_tokens(record: BranchCodeRecord) -> set[str]:
    return set(location_tokens(branch_record_text(record)))


def address_match_tokens(value: Any) -> list[str]:
    return [token for token in location_tokens(_text(value)) if token not in ADDRESS_GENERIC_TOKENS]


def address_detail_tokens(value: Any, recipient_location: str) -> list[str]:
    location_keys = name_keys_for_server(recipient_location)
    return [
        token
        for token in address_match_tokens(value)
        if normalize_location_key(token) not in location_keys
    ]


def branch_record_matches_location(record: BranchCodeRecord, recipient_location: str) -> bool:
    """Filial nomi shu hududning nomlaridan biriga to'g'ri kelsa - True.

    Zaxira tekshiruv: branch_codes.xlsx ga `location_data.py` da yo'q yangi ofis
    qo'shilgan bo'lsa ham, nomi tuman nomiga mos kelsa filial topiladi.
    """

    keys = name_keys_for_server(recipient_location)
    if not keys:
        return False
    if normalize_location_key(record.name) in keys:
        return True
    words = location_words(record.name)
    for size in range(1, min(3, len(words)) + 1):
        for start in range(len(words) - size + 1):
            if "".join(words[start : start + size]) in keys:
                return True
    return False


def branch_record_score(record: BranchCodeRecord, address: Any) -> int:
    """Manzildagi so'zlar filial nomi/manziliga qanchalik mos kelishini baholaydi."""

    record_tokens = branch_record_tokens(record)
    score = 0
    for token in address_match_tokens(address):
        if token in record_tokens:
            score += 12
        elif get_close_matches(token, list(record_tokens), n=1, cutoff=0.85):
            score += 6
    return score


def central_branch_score(record: BranchCodeRecord, recipient_location: str) -> int:
    record_tokens = branch_record_tokens(record)
    keys = name_keys_for_server(recipient_location)
    score = 0

    if normalize_location_key(record.name) in keys:
        score += 40

    score += sum(25 for token in record_tokens if token in CENTRAL_BRANCH_TOKENS)

    if "emu" in record.address.lower():
        score += 8

    return score


def best_branch_record(records: list[BranchCodeRecord], address: Any) -> BranchCodeRecord | None:
    if not records:
        return None
    if len(records) == 1:
        return records[0]
    return max(records, key=lambda record: branch_record_score(record, address))


def central_branch_record(recipient_location: str, records: list[BranchCodeRecord]) -> BranchCodeRecord | None:
    if not records:
        return None
    if len(records) == 1:
        return records[0]
    return max(records, key=lambda record: (central_branch_score(record, recipient_location), record.code))


def branch_records_for_names(names: tuple[str, ...]) -> list[BranchCodeRecord]:
    by_name = branch_records_by_name()
    found: list[BranchCodeRecord] = []
    for name in names:
        record = by_name.get(normalize_location_key(name))
        if record is not None and record not in found:
            found.append(record)
    return found


DEFAULT_BRANCH_CODE_BY_LOCATION = {
    "Самарканд": "37",
}


def default_branch_record(recipient_location: str, records: list[BranchCodeRecord]) -> BranchCodeRecord | None:
    code = DEFAULT_BRANCH_CODE_BY_LOCATION.get(recipient_location)
    if not code:
        return None
    return next((record for record in records if record.code == code), None)


def branch_code_for_address(recipient_location: str, address: Any) -> tuple[str, str]:
    """Hudud va manzil bo'yicha ДО ОФИСА uchun ichki filial kodini qaytaradi.

    Natija: (kod, izoh). Izoh bo'sh bo'lmasa - qatorni tekshirish kerak.
    """

    if not recipient_location:
        return "", ""

    records = branch_records_for_names(offices_for_server(recipient_location))
    fallback_note = ""
    if not records:
        records = [
            record
            for record in load_branch_code_records()
            if branch_record_matches_location(record, recipient_location)
        ]
    if not records:
        region_records = branch_records_for_names(region_offices_for_server(recipient_location))
        if region_records:
            records = region_records
            fallback_note = (
                f"aniq filial topilmadi, {region_center_server(recipient_location)} markaziy filiali tanlandi"
            )

    detail_tokens = address_detail_tokens(address, recipient_location)
    record = None
    if not detail_tokens:
        record = default_branch_record(recipient_location, records)
        if record is None:
            record = central_branch_record(recipient_location, records)
    if record is None:
        record = best_branch_record(records, address)
    if record is not None:
        note = fallback_note
        if fallback_note and address_match_tokens(address):
            note = f"{fallback_note} ({_text(address)})"
        return record.code, note

    return "", f"aniq filial topilmadi ({_text(address) or recipient_location})"
