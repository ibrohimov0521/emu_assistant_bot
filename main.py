import asyncio
import base64
import io
import json
import logging
import os
import re
import shutil
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime
from difflib import get_close_matches
from copy import copy
from pathlib import Path
from typing import Any

from aiogram import Bot, Dispatcher, F
from aiogram.exceptions import TelegramRetryAfter
from aiogram.filters import Command
from aiogram.types import (
    BotCommand,
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook, load_workbook

from branch_codes import branch_code_for_address
from locations import (
    SERVER_LOCATIONS,
    normalize_location_key,
    resolve_location,
    resolve_server_location,
)

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
DEFAULT_ADMIN_IDS = {6388458077}
ENV_ADMIN_IDS = {
    int(part.strip())
    for part in os.getenv("ADMIN_IDS", "").split(",")
    if part.strip().isdigit()
}
ADMIN_IDS = DEFAULT_ADMIN_IDS | ENV_ADMIN_IDS

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
TEMPLATE_DIR = BASE_DIR / "templates"
TEMPLATE_PATH = TEMPLATE_DIR / "yangi_shablon.xlsx"
PHYSICAL_EXCEL_PATH = DATA_DIR / "customers.xlsx"
LEGAL_EXCEL_PATH = DATA_DIR / "customers_legal.xlsx"
EXCEL_PATH = PHYSICAL_EXCEL_PATH
APPROVED_USERS_PATH = DATA_DIR / "approved_users.json"
ACCESS_REQUESTS_PATH = DATA_DIR / "access_requests.json"
EMU_DATABASE_PATH = DATA_DIR / "emu_database.json"

HEADERS = [
    "Номер",
    "Компания-получатель",
    "ФИО получателя",
    "Адрес получателя",
    "Телефон получателя",
    "Шифр клиента",
    "Масса посылки",
    "Поручение",
    "Количество мест",
    "Режим",
    "Компания-отправитель",
    "ФИО отправителя",
    "Адрес отправителя",
    "Телефон отправителя",
    "Город-отправитель",
    "Город-получатель",
    "Оплата получателем",
]

EXCEL_COLUMN_COUNT = len(HEADERS)
LEGAL_HEADERS = HEADERS[:10] + HEADERS[15:17]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger(__name__)

excel_lock = asyncio.Lock()
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
sender_sessions: dict[int, dict[str, str]] = {}
setup_states: dict[int, dict[str, Any]] = {}
reply_locks: dict[int, asyncio.Lock] = {}
last_success_notice_at: dict[int, float] = {}
batch_states: dict[int, "BatchState"] = {}
access_request_sent_at: dict[int, float] = {}
service_states: dict[int, dict[str, Any]] = {}
emu_api_cache: dict[str, tuple[float, Any]] = {}
emu_database_cache: dict[str, Any] | None = None
emu_database_lock = asyncio.Lock()

SUCCESS_NOTICE_INTERVAL_SECONDS = 12
BATCH_IDLE_SECONDS = 3
BATCH_CONCURRENCY = max(1, int(os.getenv("BATCH_CONCURRENCY", "10")))
BATCH_PROGRESS_EDIT_INTERVAL_SECONDS = 1.5
ACCESS_REQUEST_INTERVAL_SECONDS = 60

MENU_COLLECT = "📥 Excel ga yig'ish"
MENU_OFFICES = "🏢 Ofislar ro'yxati"
MENU_CALCULATOR = "🧮 Kalkulyator"
MENU_AI_ASSISTANT = "🤖 AI yordamchi"
MENU_ARCHIVE = "🗂 Arxiv"
MENU_SETTINGS = "⚙️ Sozlamalar"
MENU_BACK = "⬅️ Orqaga"
MENU_LEGAL = "🏢 Yuridik mijoz"
MENU_PHYSICAL = "👤 ФИЗ ЛИЦО"
MENU_EXCEL_FILE = "📊 Excel fayl"
MENU_TEMPLATE_FILE = "📄 Shablon"
MENU_CURRENT_TEMPLATES = "📑 Joriy shablonlar"
MENU_CLEAR = "🧹 Ro'yxatni tozalash"
MENU_RESET_SETUP = "✏️ Jo'natuvchi sozlamalari"
MENU_ACCESS_STATUS = "🔐 Ruxsat holati"
MENU_REFRESH_EMU_DB = "🔄 EMU bazani yangilash"
MENU_EMU_DB_STATUS = "📚 EMU baza holati"
MENU_NEXT_PAGE = "➡️ Keyingi"
MENU_PREV_PAGE = "⬅️ Oldingi"
MENU_SEARCH = "🔎 Qidirish"
MENU_CANCEL = "❌ Bekor qilish"
MENU_SKIP = "⏭️ O'tkazib yuborish"
MENU_DO_OFFICE = "🏢 ДО ОФИСА"
MENU_TO_HOME = "🏠 НА ДОМ"
MENU_TEXTS = {
    MENU_COLLECT,
    MENU_OFFICES,
    MENU_CALCULATOR,
    MENU_AI_ASSISTANT,
    MENU_ARCHIVE,
    MENU_SETTINGS,
    MENU_BACK,
    MENU_LEGAL,
    MENU_PHYSICAL,
    MENU_EXCEL_FILE,
    MENU_TEMPLATE_FILE,
    MENU_CURRENT_TEMPLATES,
    MENU_CLEAR,
    MENU_RESET_SETUP,
    MENU_ACCESS_STATUS,
    MENU_REFRESH_EMU_DB,
    MENU_EMU_DB_STATUS,
    "Excel ga yig'ish",
    "Ofislar ro'yxati",
    "Kalkulyator",
    "AI yordamchi",
    "Arxiv",
    "Sozlamalar",
    "Orqaga",
    "Yuridik mijoz",
    "ФИЗ ЛИЦО",
    "Excel fayl",
    "Shablon",
    "Joriy shablonlar",
    "Ro'yxatni tozalash",
    "Jo'natuvchi sozlamalari",
    "Ruxsat holati",
    "EMU bazani yangilash",
    "EMU baza holati",
}

CLIENT_TYPE_LEGAL = "legal"
CLIENT_TYPE_PHYSICAL = "physical"
EMU_API_BASE_URL = "https://apiv1.emu.uz"
TASHKENT_REGION_ID = 13
TASHKENT_CITY_ID = 198
EMU_CACHE_TTL_SECONDS = 3600
EMU_DB_REFRESH_INTERVAL_SECONDS = int(os.getenv("EMU_DB_REFRESH_INTERVAL_DAYS", "30")) * 24 * 60 * 60
OFFICES_PAGE_SIZE = 12

MENU_ALIASES = {
    "Excel ga yig'ish": MENU_COLLECT,
    "Ofislar ro'yxati": MENU_OFFICES,
    "Kalkulyator": MENU_CALCULATOR,
    "AI yordamchi": MENU_AI_ASSISTANT,
    "Arxiv": MENU_ARCHIVE,
    "Sozlamalar": MENU_SETTINGS,
    "Orqaga": MENU_BACK,
    "Yuridik mijoz": MENU_LEGAL,
    "ФИЗ ЛИЦО": MENU_PHYSICAL,
    "Excel fayl": MENU_EXCEL_FILE,
    "Shablon": MENU_TEMPLATE_FILE,
    "Joriy shablonlar": MENU_CURRENT_TEMPLATES,
    "Ro'yxatni tozalash": MENU_CLEAR,
    "Jo'natuvchi sozlamalari": MENU_RESET_SETUP,
    "Ruxsat holati": MENU_ACCESS_STATUS,
    "EMU bazani yangilash": MENU_REFRESH_EMU_DB,
    "EMU baza holati": MENU_EMU_DB_STATUS,
}


@dataclass
class BatchItem:
    kind: str
    message: Message
    text: str = ""
    file_id: str = ""
    mime_type: str = ""
    file_name: str = ""
    bot: Bot | None = None


@dataclass
class BatchState:
    items: list[BatchItem] = field(default_factory=list)
    task: asyncio.Task | None = None
    last_added_at: float = 0


def load_approved_user_ids() -> set[int]:
    if not APPROVED_USERS_PATH.exists():
        return set()
    try:
        data = json.loads(APPROVED_USERS_PATH.read_text(encoding="utf-8"))
    except Exception as error:
        logger.warning("Approved users file could not be read: %s", error)
        return set()
    if not isinstance(data, list):
        return set()
    return {int(item) for item in data if str(item).isdigit()}


def save_approved_user_ids(user_ids: set[int]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    APPROVED_USERS_PATH.write_text(
        json.dumps(sorted(user_ids), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


approved_user_ids: set[int] = load_approved_user_ids()


def load_access_requests() -> dict[str, dict[str, Any]]:
    if not ACCESS_REQUESTS_PATH.exists():
        return {}
    try:
        data = json.loads(ACCESS_REQUESTS_PATH.read_text(encoding="utf-8"))
    except Exception as error:
        logger.warning("Access requests file could not be read: %s", error)
        return {}
    return data if isinstance(data, dict) else {}


def save_access_requests(requests: dict[str, dict[str, Any]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    ACCESS_REQUESTS_PATH.write_text(
        json.dumps(requests, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


access_requests: dict[str, dict[str, Any]] = load_access_requests()


def empty_emu_database() -> dict[str, Any]:
    return {
        "updated_at": 0,
        "regions": [],
        "cities": [],
        "branches": [],
        "calculator_cache": {},
    }


def load_emu_database() -> dict[str, Any]:
    global emu_database_cache
    if emu_database_cache is not None:
        return emu_database_cache

    if not EMU_DATABASE_PATH.exists():
        emu_database_cache = empty_emu_database()
        return emu_database_cache

    try:
        data = json.loads(EMU_DATABASE_PATH.read_text(encoding="utf-8"))
    except Exception as error:
        logger.warning("EMU database could not be read: %s", error)
        data = {}

    database = empty_emu_database()
    if isinstance(data, dict):
        database.update({key: value for key, value in data.items() if key in database})
    if not isinstance(database.get("calculator_cache"), dict):
        database["calculator_cache"] = {}
    emu_database_cache = database
    return database


def save_emu_database(database: dict[str, Any]) -> None:
    global emu_database_cache
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    temp_path = EMU_DATABASE_PATH.with_suffix(".tmp")
    temp_path.write_text(json.dumps(database, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(EMU_DATABASE_PATH)
    emu_database_cache = database


def emu_database_age_seconds() -> float:
    updated_at = float(load_emu_database().get("updated_at") or 0)
    return max(0.0, time.time() - updated_at) if updated_at else float("inf")


def emu_database_has_core_data() -> bool:
    database = load_emu_database()
    return bool(database.get("regions") and database.get("cities") and database.get("branches"))


def emu_database_is_fresh() -> bool:
    return emu_database_has_core_data() and emu_database_age_seconds() < EMU_DB_REFRESH_INTERVAL_SECONDS


def int_value(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def region_id_from_item(item: dict[str, Any]) -> int:
    for key in ("_region_id", "region_id"):
        value = int_value(item.get(key))
        if value:
            return value
    region = item.get("region")
    if isinstance(region, dict):
        return int_value(region.get("id"))
    return 0


def city_id_from_item(item: dict[str, Any]) -> int:
    for key in ("city_id", "_city_id"):
        value = int_value(item.get(key))
        if value:
            return value
    city = item.get("city")
    if isinstance(city, dict):
        return int_value(city.get("id"))
    return 0


def calculator_cache_key(sender_city_id: int, receiver_city_id: int, weight: float, service_id: int | None) -> str:
    return json.dumps(
        {
            "sender_city_id": sender_city_id,
            "receiver_city_id": receiver_city_id,
            "weight": round(float(weight), 3),
            "service_id": service_id,
        },
        sort_keys=True,
    )


def emu_cache_key(path: str, params: dict[str, Any] | None = None) -> str:
    clean_params = {
        key: value
        for key, value in (params or {}).items()
        if value not in (None, "")
    }
    return f"{path}?{urllib.parse.urlencode(sorted(clean_params.items()))}"


def emu_api_get(path: str, params: dict[str, Any] | None = None) -> Any:
    key = emu_cache_key(path, params)
    cached = emu_api_cache.get(key)
    now = time.time()
    if cached and now - cached[0] < EMU_CACHE_TTL_SECONDS:
        return cached[1]

    query = urllib.parse.urlencode(
        {
            name: value
            for name, value in (params or {}).items()
            if value not in (None, "")
        }
    )
    url = f"{EMU_API_BASE_URL}{path}{'?' + query if query else ''}"
    request = urllib.request.Request(
        url,
        headers={
            "Accept": "application/json",
            "Accept-Language": "uz",
            "User-Agent": "EMU Assistant Bot",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        data = json.loads(response.read().decode("utf-8"))
    emu_api_cache[key] = (now, data)
    return data


def emu_api_post(path: str, payload: dict[str, Any], params: dict[str, Any] | None = None) -> Any:
    query = urllib.parse.urlencode(
        {
            name: value
            for name, value in (params or {}).items()
            if value not in (None, "")
        }
    )
    url = f"{EMU_API_BASE_URL}{path}{'?' + query if query else ''}"
    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Accept": "application/json",
            "Accept-Language": "uz",
            "Content-Type": "application/json",
            "User-Agent": "EMU Assistant Bot",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def refresh_emu_database_sync() -> dict[str, Any]:
    old_database = load_emu_database()
    emu_api_cache.clear()

    regions = emu_api_get("/api/v1/regions")
    cities = emu_api_get("/api/v1/cities")
    branches: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for region in regions:
        region_id = int_value(region.get("id"))
        if not region_id:
            continue
        region_name = localized_name(region)
        for branch in emu_api_get("/api/v1/branches", {"region_id": region_id}):
            branch_id = clean_text(branch.get("id")) or f"{region_id}:{clean_text(branch.get('name'))}"
            if branch_id in seen_ids:
                continue
            seen_ids.add(branch_id)
            enriched = dict(branch)
            enriched["_region_id"] = region_id
            enriched["_region_name"] = region_name
            branches.append(enriched)

    database = {
        "updated_at": time.time(),
        "regions": regions,
        "cities": cities,
        "branches": branches,
        "calculator_cache": old_database.get("calculator_cache") if isinstance(old_database.get("calculator_cache"), dict) else {},
    }
    save_emu_database(database)
    return database


async def refresh_emu_database() -> dict[str, Any]:
    async with emu_database_lock:
        return await asyncio.to_thread(refresh_emu_database_sync)


async def ensure_emu_database() -> dict[str, Any]:
    if emu_database_has_core_data():
        return load_emu_database()
    return await refresh_emu_database()


async def get_emu_regions() -> list[dict[str, Any]]:
    database = await ensure_emu_database()
    return sort_emu_regions(list(database.get("regions") or []))


async def get_emu_cities(region_id: int | None = None) -> list[dict[str, Any]]:
    database = await ensure_emu_database()
    cities = list(database.get("cities") or [])
    if region_id:
        cities = [city for city in cities if region_id_from_item(city) == int(region_id)]
    return sort_emu_cities(cities)


async def get_emu_branches(region_id: int | None = None, city_id: int | None = None) -> list[dict[str, Any]]:
    database = await ensure_emu_database()
    branches = list(database.get("branches") or [])
    if region_id:
        branches = [branch for branch in branches if region_id_from_item(branch) == int(region_id)]
    if city_id:
        city_filtered = [branch for branch in branches if city_id_from_item(branch) == int(city_id)]
        if int(city_id) == TASHKENT_CITY_ID and not city_filtered:
            city_filtered = [branch for branch in branches if region_id_from_item(branch) == TASHKENT_REGION_ID]
        if not city_filtered:
            cities = list(database.get("cities") or [])
            city = next((item for item in cities if int_value(item.get("id")) == int(city_id)), None)
            if city:
                city_keys = city_search_keys(city)
                city_filtered = [
                    branch
                    for branch in branches
                    if city_keys
                    & {
                        compact_region_key(clean_text(branch.get("city_name"))),
                        compact_region_key(clean_text(branch.get("name"))),
                        compact_region_key(clean_text(branch.get("address"))),
                    }
                ]
        branches = city_filtered
    return sort_emu_branches(branches)


async def get_all_emu_branches() -> list[dict[str, Any]]:
    database = await ensure_emu_database()
    return sort_emu_branches(list(database.get("branches") or []))


async def calculate_emu_delivery(
    sender_city_id: int,
    receiver_city_id: int,
    weight: float,
    service_id: int | None = None,
) -> dict[str, Any]:
    database = await ensure_emu_database()
    cache_key = calculator_cache_key(sender_city_id, receiver_city_id, weight, service_id)
    calculator_cache = database.setdefault("calculator_cache", {})
    cached_result = calculator_cache.get(cache_key) if isinstance(calculator_cache, dict) else None
    if isinstance(cached_result, dict) and cached_result.get("result"):
        return cached_result["result"]

    payload = {
        "sender_city_id": sender_city_id,
        "receiver_city_id": receiver_city_id,
        "service": service_id,
        "packages": [
            {
                "mass": weight,
                "length": None,
                "width": None,
                "height": None,
            }
        ],
    }
    result = await asyncio.to_thread(emu_api_post, "/api/v1/calculator", payload, {"platform": "app"})
    async with emu_database_lock:
        database = load_emu_database()
        calculator_cache = database.setdefault("calculator_cache", {})
        calculator_cache[cache_key] = {"updated_at": time.time(), "result": result}
        save_emu_database(database)
    return result


def localized_name(item: dict[str, Any], locale: str = "UZ") -> str:
    i18n = item.get("i18n_name")
    if isinstance(i18n, dict):
        return clean_text(i18n.get(locale)) or clean_text(i18n.get("UZ")) or clean_text(item.get("name"))
    return clean_text(item.get("name"))


def alpha_key(value: str) -> str:
    text = clean_text(value).casefold()
    replacements = {
        "o‘": "o",
        "o'": "o",
        "g‘": "g",
        "g'": "g",
        "ʻ": "",
        "‘": "",
        "ʼ": "",
        "`": "",
        "sh": "s~h",
        "ch": "c~h",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return re.sub(r"\s+", " ", text).strip()


def region_sort_key(region: dict[str, Any]) -> tuple[int, str]:
    name = localized_name(region)
    normalized = alpha_key(name)
    if int_value(region.get("id")) == TASHKENT_REGION_ID or normalized == "toshkent":
        return (0, normalized)
    return (1, normalized)


def city_sort_key(city: dict[str, Any]) -> tuple[int, str]:
    if int_value(city.get("id")) == TASHKENT_CITY_ID:
        return (0, "")
    return (1, alpha_key(localized_name(city)))


def branch_sort_key(branch: dict[str, Any]) -> tuple[str, str, str]:
    return (
        alpha_key(clean_text(branch.get("name"))),
        alpha_key(clean_text(branch.get("city_name"))),
        alpha_key(clean_text(branch.get("address"))),
    )


def sort_emu_regions(regions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(regions, key=region_sort_key)


def sort_emu_cities(cities: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(cities, key=city_sort_key)


def sort_emu_branches(branches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(branches, key=branch_sort_key)


def chunked(items: list[Any], size: int) -> list[list[Any]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def reply_keyboard(labels: list[str], row_size: int = 2, add_back: bool = True) -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(text=label) for label in row]
        for row in chunked(labels, row_size)
    ]
    if add_back:
        rows.append([KeyboardButton(text=MENU_BACK)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def remember_options(state: dict[str, Any], items: list[dict[str, Any]], locale: str = "UZ") -> list[str]:
    labels: list[str] = []
    options: dict[str, int] = {}
    for item in items:
        label = localized_name(item, locale)
        if not label:
            continue
        labels.append(label)
        options[label] = int(item.get("id") or 0)
    state["options"] = options
    return labels


def selected_option_id(state: dict[str, Any], text: str) -> int | None:
    options = state.get("options") or {}
    if text in options:
        return int(options[text])
    normalized = clean_text(text).casefold()
    for label, value in options.items():
        if clean_text(label).casefold() == normalized:
            return int(value)
    return None


def region_reply_keyboard(regions: list[dict[str, Any]], state: dict[str, Any]) -> ReplyKeyboardMarkup:
    return reply_keyboard(remember_options(state, regions), row_size=2)


def city_reply_keyboard(cities: list[dict[str, Any]], state: dict[str, Any]) -> ReplyKeyboardMarkup:
    return reply_keyboard(remember_options(state, cities), row_size=2)


def calculator_service_reply_keyboard() -> ReplyKeyboardMarkup:
    return reply_keyboard([MENU_DO_OFFICE, MENU_TO_HOME], row_size=2)


def offices_page_keyboard(page: int, total: int) -> ReplyKeyboardMarkup:
    labels: list[str] = []
    if page > 0:
        labels.append(MENU_PREV_PAGE)
    if (page + 1) * OFFICES_PAGE_SIZE < total:
        labels.append(MENU_NEXT_PAGE)
    labels.append(MENU_SEARCH)
    return reply_keyboard(labels, row_size=2)


def region_keyboard(regions: list[dict[str, Any]], prefix: str) -> InlineKeyboardMarkup:
    buttons = [
        InlineKeyboardButton(text=localized_name(region), callback_data=f"{prefix}:{region['id']}")
        for region in regions
    ]
    rows = chunked(buttons, 2)
    rows.append([InlineKeyboardButton(text=MENU_BACK, callback_data="emu:back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def city_keyboard(cities: list[dict[str, Any]], prefix: str) -> InlineKeyboardMarkup:
    buttons = [
        InlineKeyboardButton(text=localized_name(city), callback_data=f"{prefix}:{city['id']}")
        for city in cities
    ]
    rows = chunked(buttons, 2)
    rows.append([InlineKeyboardButton(text=MENU_BACK, callback_data="emu:back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def service_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="ДО ОФИСА", callback_data="emu:calc_service:1"),
                InlineKeyboardButton(text="НА ДОМ", callback_data="emu:calc_service:3"),
            ],
            [InlineKeyboardButton(text=MENU_BACK, callback_data="emu:back")],
        ]
    )


def format_branch_card(branch: dict[str, Any], index: int) -> str:
    schedule = branch.get("work_schedule") or []
    active_days = [day for day in schedule if day.get("is_active")]
    work_time = ""
    if active_days:
        first = active_days[0]
        work_time = f"{clean_text(first.get('start_time'))[:5]}-{clean_text(first.get('end_time'))[:5]}"
    open_status = "🟢 ochiq" if branch.get("is_open_now") else "🔴 yopiq"
    phone = clean_text(branch.get("phone")) or "ko'rsatilmagan"
    address_parts = [clean_text(branch.get("address")), clean_text(branch.get("address_ref"))]
    address = " | ".join(part for part in address_parts if part)
    region = clean_text(branch.get("region_name")) or clean_text(branch.get("_region_name"))
    city = clean_text(branch.get("city_name"))
    address_text = address or "ko'rsatilmagan"
    work_text = work_time or "ko'rsatilmagan"
    return (
        f"{index}. 🏢 {clean_text(branch.get('name'))}\n"
        f"📍 Hudud: {region}, {city}\n"
        f"🧭 Manzil: {address_text}\n"
        f"📞 Tel: {phone}\n"
        f"🕘 Ish vaqti: {work_text}\n"
        f"🚪 Holat: {open_status}"
    )


def format_branches_list(branches: list[dict[str, Any]], title: str, limit: int = OFFICES_PAGE_SIZE) -> str:
    if not branches:
        return f"{title}\n\nBu hudud uchun ofis topilmadi."

    visible = branches[:limit]
    lines = [f"🏢 {title}", f"📌 Jami: {len(branches)} ta ofis", "🔤 Tartib: alfavit bo'yicha"]
    lines.extend(format_branch_card(branch, index) for index, branch in enumerate(visible, start=1))
    if len(branches) > limit:
        lines.append(f"\n➡️ Yana {len(branches) - limit} ta ofis bor. Aniq tuman bo'yicha qidirsak, ro'yxat qisqaradi.")
    return "\n\n".join(lines)


def format_branches_page(branches: list[dict[str, Any]], title: str, page: int = 0) -> str:
    if not branches:
        return f"🏢 {title}\n\n😕 Bu hudud uchun ofis topilmadi."

    total = len(branches)
    page_count = max(1, (total + OFFICES_PAGE_SIZE - 1) // OFFICES_PAGE_SIZE)
    page = max(0, min(page, page_count - 1))
    start = page * OFFICES_PAGE_SIZE
    visible = branches[start : start + OFFICES_PAGE_SIZE]

    lines = [
        f"🏢 {title}",
        f"📌 Jami: {total} ta ofis",
        f"📄 Sahifa: {page + 1}/{page_count}",
        "🔤 Tartib: alfavit bo'yicha",
    ]
    lines.extend(
        format_branch_card(branch, index)
        for index, branch in enumerate(visible, start=start + 1)
    )
    if page + 1 < page_count:
        lines.append(f"\n➡️ Yana {total - (start + len(visible))} ta ofis bor. Pastdan {MENU_NEXT_PAGE} ni bosing.")
    return "\n\n".join(lines)


def filter_branches(branches: list[dict[str, Any]], query: str) -> list[dict[str, Any]]:
    query = clean_text(query).casefold()
    if not query:
        return branches
    return [
        branch
        for branch in branches
        if query
        in " ".join(
            [
                clean_text(branch.get("name")),
                clean_text(branch.get("address")),
                clean_text(branch.get("address_ref")),
                clean_text(branch.get("city_name")),
                clean_text(branch.get("region_name")),
                clean_text(branch.get("phone")),
            ]
        ).casefold()
    ]


OFFICE_QUERY_STOP_WORDS = {
    "ofis",
    "ofislar",
    "ofisbor",
    "ofisi",
    "filial",
    "filiallar",
    "nechta",
    "qancha",
    "qayerda",
    "qanaqa",
    "bormi",
    "bor",
    "yoq",
    "yo'q",
    "office",
    "branch",
}


def compact_region_key(value: str) -> str:
    text = clean_text(value).casefold()
    text = re.sub(
        r"\b(viloyati|viloyat|область|обл|respublikasi|respublika|республика|город|shahri|shahar)\b",
        " ",
        text,
        flags=re.IGNORECASE,
    )
    return normalize_location_key(text)


def branch_region_keys(branch: dict[str, Any]) -> set[str]:
    keys: set[str] = set()
    for value in [
        clean_text(branch.get("_region_name")),
        clean_text(branch.get("region_name")),
        clean_text(branch.get("city_name")),
    ]:
        key = compact_region_key(value)
        if key:
            keys.add(key)
    return keys


def branch_search_key(branch: dict[str, Any]) -> str:
    values = [
        clean_text(branch.get("name")),
        clean_text(branch.get("address")),
        clean_text(branch.get("address_ref")),
        clean_text(branch.get("city_name")),
        clean_text(branch.get("region_name")),
        clean_text(branch.get("_region_name")),
    ]
    return normalize_location_key(" ".join(values))


def office_query_words(question: str) -> set[str]:
    words = set()
    for word in re.findall(r"[\w'`‘’.-]+", question.casefold()):
        normalized = strip_place_suffix(word)
        if len(normalized) >= 3 and normalized not in OFFICE_QUERY_STOP_WORDS:
            words.add(normalized)
    return words


def find_matching_branches_for_question(branches: list[dict[str, Any]], question: str) -> list[dict[str, Any]]:
    question_key = normalize_location_key(question)
    words = office_query_words(question)
    if not question_key and not words:
        return []

    region_matches = [
        branch
        for branch in branches
        if any(region_key and region_key in question_key for region_key in branch_region_keys(branch))
    ]
    if region_matches:
        return region_matches

    matching = []
    for branch in branches:
        branch_key = branch_search_key(branch)
        if any(word and (word in branch_key or branch_key in word) for word in words):
            matching.append(branch)
    return matching


def format_office_count_answer(branches: list[dict[str, Any]]) -> str:
    if not branches:
        return "Bu hudud bo'yicha ofis topilmadi. Viloyat yoki tuman nomini aniqroq yozing."

    names = [clean_text(branch.get("name")) for branch in branches if clean_text(branch.get("name"))]
    unique_names = list(dict.fromkeys(names))
    region_name = clean_text(branches[0].get("_region_name")) or "tanlangan hudud"
    lines = [f"{region_name} bo'yicha {len(branches)} ta ofis bor."]
    if unique_names:
        lines.append("")
        for index, name in enumerate(unique_names[:20], start=1):
            lines.append(f"{index}. {name}")
        if len(unique_names) > 20:
            lines.append(f"... yana {len(unique_names) - 20} ta ofis bor.")
    return "\n".join(lines)


PLACE_SUFFIXES = ("gacha", "dan", "danmi", "ga", "da", "ni", "ning", "mi")


def city_search_keys(city: dict[str, Any]) -> set[str]:
    values = [
        clean_text(city.get("name")),
        clean_text(city.get("extra_name")),
        localized_name(city, "UZ"),
        localized_name(city, "RU"),
        localized_name(city, "EN"),
    ]
    keys: set[str] = set()
    for value in values:
        key = compact_region_key(value)
        if key:
            keys.add(key)
        simplified = re.sub(
            r"\b(tumani|shahri|shahar|district|city|район|город|г)\b",
            " ",
            value.casefold(),
            flags=re.IGNORECASE,
        )
        simplified_key = compact_region_key(simplified)
        if simplified_key:
            keys.add(simplified_key)
    return keys


def strip_place_suffix(value: str) -> str:
    key = normalize_location_key(value)
    for suffix in PLACE_SUFFIXES:
        if key.endswith(suffix) and len(key) > len(suffix) + 2:
            return key[: -len(suffix)]
    return key


def find_city_by_text(text: str, cities: list[dict[str, Any]]) -> dict[str, Any] | None:
    query_key = normalize_location_key(text)
    if not query_key:
        return None

    exact_matches: list[tuple[int, dict[str, Any]]] = []
    key_to_city: dict[str, dict[str, Any]] = {}
    for city in cities:
        for key in city_search_keys(city):
            key_to_city.setdefault(key, city)
            if key and key in query_key:
                exact_matches.append((len(key), city))

    if exact_matches:
        exact_matches.sort(key=lambda item: item[0], reverse=True)
        return exact_matches[0][1]

    words = [strip_place_suffix(word) for word in re.findall(r"[\w'`‘’.-]+", text.casefold())]
    words = [word for word in words if len(word) >= 4]
    all_keys = list(key_to_city)
    for word in words:
        matches = get_close_matches(word, all_keys, n=1, cutoff=0.82)
        if matches:
            return key_to_city[matches[0]]
    return None


def parse_weight_from_question(question: str) -> float | None:
    match = re.search(r"(?i)(\d+(?:[.,]\d+)?)\s*(?:kg|кг|kilogram|кило)\b", question)
    if not match:
        return None
    try:
        return float(match.group(1).replace(",", "."))
    except ValueError:
        return None


def ai_service_id_from_question(question: str, sender: dict[str, str] | None = None) -> int:
    lowered = question.casefold()
    if any(word in lowered for word in ["uyga", "uygacha", "na dom", "на дом", "дом"]):
        return 3
    if any(word in lowered for word in ["ofis", "ofisgacha", "pochta", "до офиса", "do ofisa"]):
        return 1
    if sender and sender.get("delivery_type") == "НА ДОМ":
        return 3
    return 1


def ai_origin_destination_text(question: str) -> tuple[str, str]:
    lowered = question.casefold()
    match = re.search(r"(.+?)\s+dan\s+(.+?)(?:\s+ga|\s+gacha|\s|$)", lowered)
    if match:
        return match.group(1), match.group(2)
    match = re.search(r"([\w'`‘’.-]+)dan\s+(.+?)(?:\s|$)", lowered)
    if match:
        return match.group(1), match.group(2)
    match = re.search(r"(.+?)\s+дан\s+(.+?)(?:\s+га|\s|$)", lowered)
    if match:
        return match.group(1), match.group(2)
    return "", question


async def calculate_from_ai_question(message: Message, question: str) -> str | None:
    weight = parse_weight_from_question(question)
    if weight is None:
        return None

    cities = await get_emu_cities()
    sender = sender_sessions.get(message.chat.id)
    origin_text, destination_text = ai_origin_destination_text(question)

    sender_city = find_city_by_text(origin_text, cities) if origin_text else None
    if sender_city is None and sender:
        sender_city = find_city_by_text(sender.get("sender_city_ru", ""), cities)
    if sender_city is None:
        return (
            "🧮 Hisoblash uchun jo'natilish nuqtasi yetishmayapti.\n\n"
            "Masalan:\n"
            "Toshkentdan Paxtaobodga 1 kg pochta qancha?"
        )

    receiver_city = find_city_by_text(destination_text, cities)
    if receiver_city is None:
        return (
            "🧮 Olish nuqtasini aniqlay olmadim.\n\n"
            "Viloyat yoki tuman nomini aniqroq yozing. Masalan:\n"
            "Toshkentdan Paxtaobodga 1 kg pochta qancha?"
        )

    service_id = ai_service_id_from_question(question, sender)
    result = await calculate_emu_delivery(
        int(sender_city["id"]),
        int(receiver_city["id"]),
        weight,
        service_id,
    )
    receiver_branches = await get_emu_branches(city_id=int(receiver_city["id"]))
    return format_calculator_result(result, service_id, receiver_branches)


def format_calculator_result(
    result: dict[str, Any],
    service_id: int,
    receiver_branches: list[dict[str, Any]],
) -> str:
    results = result.get("results") if isinstance(result, dict) else []
    selected = None
    for item in results or []:
        if int(item.get("service_id") or 0) == service_id:
            selected = item
            break
    if selected is None and results:
        selected = results[0]

    if not selected:
        price_text = "Narx topilmadi"
    else:
        price = selected.get("price")
        currency = selected.get("currency") or "UZS"
        days = ""
        if selected.get("min_delivery_days") and selected.get("max_delivery_days"):
            days = f"\nMuddat: {selected['min_delivery_days']}-{selected['max_delivery_days']} kun"
        price_value = f"{float(price):,.0f}".replace(",", " ") if price is not None else "topilmadi"
        price_text = (
            f"Xizmat: {clean_text(selected.get('service_name'))}\n"
            f"Narx: {price_value} {currency}"
            + days
        )

    branch_lines = []
    if receiver_branches:
        branch_lines.append("Mavjud ofislar:")
        for branch in receiver_branches[:5]:
            branch_lines.append(f"- {clean_text(branch.get('name'))}: {clean_text(branch.get('address'))}")
        if len(receiver_branches) > 5:
            branch_lines.append(f"... yana {len(receiver_branches) - 5} ta ofis bor")
    else:
        branch_lines.append("Bu tuman/shahar uchun ofis topilmadi.")

    return "Hisob-kitob natijasi:\n\n" + price_text + "\n\n" + "\n".join(branch_lines)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def has_bot_access(user_id: int) -> bool:
    if not ADMIN_IDS:
        return True
    return is_admin(user_id) or user_id in approved_user_ids


def main_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=MENU_COLLECT)],
            [KeyboardButton(text=MENU_OFFICES), KeyboardButton(text=MENU_CALCULATOR)],
            [KeyboardButton(text=MENU_AI_ASSISTANT)],
            [KeyboardButton(text=MENU_ARCHIVE), KeyboardButton(text=MENU_SETTINGS)],
        ],
        resize_keyboard=True,
    )


def collect_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=MENU_LEGAL), KeyboardButton(text=MENU_PHYSICAL)],
            [KeyboardButton(text=MENU_BACK)],
        ],
        resize_keyboard=True,
    )


def archive_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=MENU_EXCEL_FILE), KeyboardButton(text=MENU_TEMPLATE_FILE)],
            [KeyboardButton(text=MENU_CLEAR)],
            [KeyboardButton(text=MENU_BACK)],
        ],
        resize_keyboard=True,
    )


def settings_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=MENU_RESET_SETUP), KeyboardButton(text=MENU_ACCESS_STATUS)],
            [KeyboardButton(text=MENU_REFRESH_EMU_DB), KeyboardButton(text=MENU_EMU_DB_STATUS)],
            [KeyboardButton(text=MENU_CURRENT_TEMPLATES)],
            [KeyboardButton(text=MENU_BACK)],
        ],
        resize_keyboard=True,
    )


def collect_active_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=MENU_EXCEL_FILE), KeyboardButton(text=MENU_TEMPLATE_FILE)],
            [KeyboardButton(text=MENU_BACK)],
        ],
        resize_keyboard=True,
    )


SETUP_STEPS = [
    (
        "sender_full_name",
        "Jo'natuvchining ism familiyasini kiriting.",
    ),
    (
        "sender_phone",
        "Jo'natuvchining telefon raqamini kiriting. Masalan: +998 90 123 45 67",
    ),
    (
        "sender_address",
        "Jo'natuvchining to'liq manzilini kiriting.",
    ),
    (
        "sender_city_ru",
        "Jo'natuvchi qaysi tuman/shahardanligini rus tilida kiriting. Masalan: Ташкент, Бухара, Шафиркан",
    ),
    (
        "cipher_prefix",
        "Jo'natmalar uchun qaytarilmaydigan shifr prefixini kiriting. Masalan: ABC",
    ),
    (
        "delivery_type",
        "Yetkazib berish turini tanlang.",
    ),
    (
        "payment_by_receiver",
        "Оплата получателем bo'ladimi?",
    ),
    (
        "parcel_weight",
        "Jo'natma og'irligini kiriting. Masalan: 1.5",
    ),
    (
        "places_count",
        "Bir mijozga nechta jo'natma bo'lishini kiriting. Masalan: 1",
    ),
]

ALLOWED_RECIPIENT_LOCATIONS = list(SERVER_LOCATIONS)

LOCATION_LIST_FOR_PROMPT = ", ".join(ALLOWED_RECIPIENT_LOCATIONS)

BUTTON_SETUP_OPTIONS = {
    "delivery_type": [
        (MENU_DO_OFFICE, "ДО ОФИСА"),
        (MENU_TO_HOME, "НА ДОМ"),
    ],
    "payment_by_receiver": [
        ("✅ Galochka qo'yilsin", "True"),
        ("⬜ Galochka qo'yilmasin", "False"),
    ],
    "cipher_prefix": [
        (MENU_SKIP, ""),
    ],
}


CUSTOMER_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "customers": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "number": {"type": "string"},
                    "full_name": {"type": "string"},
                    "phone": {"type": "string"},
                    "address": {"type": "string"},
                    "recipient_region_ru": {"type": "string"},
                    "note": {"type": "string"},
                    "needs_review": {"type": "string"},
                },
                "required": [
                    "number",
                    "full_name",
                    "phone",
                    "address",
                    "recipient_region_ru",
                    "note",
                    "needs_review",
                ],
            },
        }
    },
    "required": ["customers"],
}


SYSTEM_PROMPT = """
Siz mijoz ma'lumotlarini ajratadigan yordamchisiz.
Matn yoki rasmda bir nechta mijoz bo'lishi mumkin. Har bir mijozni alohida obyekt qiling.

Ajratiladigan maydonlar:
- number: asl tartib raqami bor bo'lsa, aks holda bo'sh string
- full_name: ism familiya bor bo'lsa
- phone: barcha telefon raqamlari asl matndagi ko'rinishida; bir nechta bo'lsa hammasini shu maydonga yozing
- address: manzil
- recipient_region_ru: oluvchining manzilidan P ustun uchun mos shahar/tuman nomini rus tilida tanlang. Faqat quyidagi ro'yxatdan bittasini yozing, boshqa format yozmang:
{location_list}
- note: boshqa foydali izohlar, noaniq yoki yo'qolmasligi kerak bo'lgan bo'laklar; telefon raqamlarini bu maydonga yozmang
- needs_review: noaniq o'qilgan, telefon raqami shubhali, rasm sifati past, yoki maydonlar aralash bo'lsa qisqa izoh

Qoidalar:
- Ma'lumot yo'q bo'lsa bo'sh string qaytaring.
- Telefon raqamlarni formatlamang, asl ko'rinishida qaytaring.
- Taxmin qilmang. Ishonchsiz joylarni needs_review maydoniga yozing.
- Ism yo'q bo'lsa "Mijoz", "Noma'lum", "Customer" kabi placeholder yozmang, full_name bo'sh string bo'lsin.
- Bitta xabarda bitta ism/manzil va bir nechta telefon raqami bo'lsa, buni bitta mijoz deb oling: hamma telefon raqamlarni phone maydoniga yozing, note maydoniga telefon yozmang.
- Faqat aniq boshqa-boshqa mijozlar bo'lsa alohida obyekt qiling.
- address maydoniga faqat yetkazish manzilini yozing. Narx, to'lov summasi, rang, tovar/vlojenie, o'lcham, kiyim turi kabi ma'lumotlarni addressga qo'shmang; ularni note maydoniga yozing.
- Masalan: "Farg'ona Mustaqillik ko'cha mehmonxona oldi, 360.000 so'm qora" bo'lsa address="Farg'ona Mustaqillik ko'cha mehmonxona oldi", note="360.000 so'm qora".
- recipient_region_ru hech qachon "Ферганская область, Учкуприкский район" kabi bo'lmasin; ro'yxatdagi "Учкуприк" kabi bitta qiymat bo'lsin.
- Agar manzilda viloyat/tuman/shahar nomi bor bo'lsa, recipient_region_ru ni bo'sh qoldirmang; ro'yxatdan eng yaqin mos qiymatni tanlang.
- "Samarqand viloyati Paxtachi tumani" bo'lsa recipient_region_ru uchun "Пахтачи" yozing.
- Tuman yoki shahar nomi viloyatdan muhimroq: "Farg'ona viloyati Oltiariq tumani" uchun "Фергана" emas, "Алтыарык" yozing.
- Lotin yozuvidagi O'zbekcha nomlarni ruscha ro'yxatga moslang: Qorako'l -> Каракуль, Qo'rg'ontepa -> Кургантепа, Bo'ka -> Бука, Tayloq -> Тайлак.
- Javob faqat schema bo'yicha bo'lsin.
""".strip().format(location_list=LOCATION_LIST_FOR_PROMPT)


def headers_for_client_type(client_type: str = CLIENT_TYPE_PHYSICAL) -> list[str]:
    return LEGAL_HEADERS if client_type == CLIENT_TYPE_LEGAL else HEADERS


def excel_path_for_client_type(client_type: str = CLIENT_TYPE_PHYSICAL) -> Path:
    return LEGAL_EXCEL_PATH if client_type == CLIENT_TYPE_LEGAL else PHYSICAL_EXCEL_PATH


def current_client_type(chat_id: int) -> str:
    sender = sender_sessions.get(chat_id) or {}
    return sender.get("client_type", CLIENT_TYPE_PHYSICAL)


def template_column_map(template_sheet: Any) -> dict[str, int]:
    return {
        clean_text(template_sheet.cell(1, col).value): col
        for col in range(1, template_sheet.max_column + 1)
        if clean_text(template_sheet.cell(1, col).value)
    }


def apply_template_header(sheet: Any, headers: list[str] | None = None) -> None:
    headers = headers or HEADERS
    if TEMPLATE_PATH.exists():
        template = load_workbook(TEMPLATE_PATH)
        try:
            template_sheet = template.active
            source_columns = template_column_map(template_sheet)
            for col, header in enumerate(headers, start=1):
                source_col = source_columns.get(header, col)
                source = template_sheet.cell(1, source_col)
                target = sheet.cell(1, col)
                target.value = header
                if source.has_style:
                    target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
                letter = target.column_letter
                source_letter = source.column_letter
                sheet.column_dimensions[letter].width = template_sheet.column_dimensions[source_letter].width
        finally:
            template.close()
        return

    for col, header in enumerate(headers, start=1):
        sheet.cell(1, col).value = header


def ensure_workbook_schema(path: Path, headers: list[str]) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        current_headers = [sheet.cell(1, col).value for col in range(1, len(headers) + 1)]

        if current_headers == headers and sheet.max_column == len(headers):
            return

        if current_headers == headers and sheet.max_column > len(headers):
            sheet.delete_cols(len(headers) + 1, sheet.max_column - len(headers))
            workbook.save(path)
            return

        first_row_has_data = any(value not in (None, "") for value in current_headers)
        if first_row_has_data:
            header_to_source = {
                clean_text(sheet.cell(1, col).value): col
                for col in range(1, sheet.max_column + 1)
                if clean_text(sheet.cell(1, col).value)
            }
            if all(header in header_to_source for header in headers):
                rows = [
                    [sheet.cell(row_index, header_to_source[header]).value for header in headers]
                    for row_index in range(2, sheet.max_row + 1)
                ]
                sheet.delete_rows(1, sheet.max_row)
                apply_template_header(sheet, headers)
                for row in rows:
                    sheet.append(row)
            else:
                sheet.insert_rows(1)
                apply_template_header(sheet, headers)
        else:
            apply_template_header(sheet, headers)

        if sheet.max_column > len(headers):
            sheet.delete_cols(len(headers) + 1, sheet.max_column - len(headers))
        workbook.save(path)
    finally:
        workbook.close()


def ensure_excel_file(client_type: str = CLIENT_TYPE_PHYSICAL) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path = excel_path_for_client_type(client_type)
    headers = headers_for_client_type(client_type)
    if path.exists():
        ensure_workbook_schema(path, headers)
        return

    if TEMPLATE_PATH.exists():
        shutil.copyfile(TEMPLATE_PATH, path)
        workbook = load_workbook(path)
        try:
            sheet = workbook.active
            apply_template_header(sheet, headers)
            if sheet.max_column > len(headers):
                sheet.delete_cols(len(headers) + 1, sheet.max_column - len(headers))
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.value = None
            workbook.save(path)
        finally:
            workbook.close()
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Шаблон"
    sheet.append(headers)
    widths = [18, 24, 19, 21, 23, 18, 13, 13, 20, 27, 25, 20, 22, 24, 22, 21, 23]
    for index, width in enumerate(widths, start=1):
        if index <= len(headers):
            sheet.column_dimensions[chr(64 + index)].width = width

    workbook.save(path)


def reset_excel_file(client_type: str = CLIENT_TYPE_PHYSICAL) -> None:
    path = excel_path_for_client_type(client_type)
    if path.exists():
        path.unlink()
    ensure_excel_file(client_type)


def normalize_phone(raw_phone: str) -> tuple[str, str]:
    raw_phone = (raw_phone or "").strip()
    if not raw_phone:
        return "", ""

    digits = re.sub(r"\D", "", raw_phone)
    review = ""

    if len(digits) == 12 and digits.startswith("998"):
        return digits, review

    if len(digits) == 9:
        return f"998{digits}", review

    if len(digits) == 10 and digits.startswith("0"):
        return f"998{digits[1:]}", review

    return raw_phone, f"Telefon noaniq: {raw_phone}"


PHONE_CANDIDATE_RE = re.compile(r"\+?\d[\d\s\-()]{7,}\d")
PRICE_RE = re.compile(
    r"(?i)(?:\b\d{1,3}(?:[ .]\d{3})+(?:[,.]\d+)?|\b\d+(?:[,.]\d+)?)\s*"
    r"(?:so['‘’`]?m|sum|som|сум|сўм|uzs)\b"
)
ADDRESS_HINT_RE = re.compile(
    r"(?i)\b("
    r"viloyat|vilo?yati|tuman|tumani|shahar|shahri|ko['‘’`]?cha|ko['‘’`]?chasi|"
    r"mahalla|mfy|massiv|mavze|daha|uy|dom|kv|xonadon|orientir|mo['‘’`]?ljal|"
    r"oldi|orqa|yonida|qarshi|ro['‘’`]?para|bekat|bozor|maktab|bog['‘’`]?cha|"
    r"г\.|город|область|район|улица|ул\.|махалля|дом|кв|ориентир|рядом|напротив"
    r")\b"
)
PARCEL_NOTE_RE = re.compile(
    r"(?i)\b("
    r"qora|oq|qizil|ko['‘’`]?k|kok|yashil|sariq|kulrang|jigarrang|pushti|"
    r"черн|бел|красн|син|зел|желт|сер|корич|роз|"
    r"kiyim|kuylak|shim|oyoq|sumka|kross|tufli|razmer|o['‘’`]?lcham|размер|"
    r"платье|брюк|обув|сумк|товар|dona|ta|шт"
    r")\b"
)


def extract_phone_candidates(*values: str) -> list[str]:
    candidates: list[str] = []
    for value in values:
        for match in PHONE_CANDIDATE_RE.findall(value or ""):
            candidate = match.strip(" -()")
            if candidate:
                candidates.append(candidate)
    return candidates


def normalize_phone_list(*values: str) -> tuple[str, str]:
    normalized_numbers: list[str] = []
    reviews: list[str] = []

    for candidate in extract_phone_candidates(*values):
        normalized, review = normalize_phone(candidate)
        if review:
            reviews.append(review)
            continue
        if normalized not in normalized_numbers:
            normalized_numbers.append(normalized)

    if normalized_numbers:
        return "; ".join(normalized_numbers), "; ".join(reviews)

    raw_phone = clean_text(values[0]) if values else ""
    if not raw_phone:
        return "", ""
    return normalize_phone(raw_phone)


def strip_phone_candidates(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = PHONE_CANDIDATE_RE.sub("", text)
    text = re.sub(r"\b(qolgan|ikkinchi|2-?chi|telefon|tel|raqamlar|raqami)\b\s*:?", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*[,;|/]\s*", " ", text)
    return re.sub(r"\s{2,}", " ", text).strip(" -,:;|/")


def clean_split_part(value: str) -> str:
    return re.sub(r"\s{2,}", " ", clean_text(value)).strip(" -,:;|/")


def looks_like_parcel_note(value: str) -> bool:
    text = clean_split_part(value)
    if not text:
        return False
    if PRICE_RE.search(text):
        return True
    return bool(PARCEL_NOTE_RE.search(text)) and not ADDRESS_HINT_RE.search(text)


def split_address_and_note(value: Any) -> tuple[str, str]:
    text = strip_phone_candidates(clean_text(value))
    if not text:
        return "", ""

    price_match = PRICE_RE.search(text)
    if price_match:
        address = clean_split_part(text[: price_match.start()])
        note = clean_split_part(text[price_match.start() :])
        return address, note

    separators = re.split(r"([,;\n]+)", text)
    parts = [clean_split_part(part) for part in separators if clean_split_part(part) and not re.fullmatch(r"[,;\n]+", part)]
    note_parts: list[str] = []
    while parts and looks_like_parcel_note(parts[-1]):
        note_parts.insert(0, parts.pop())

    address = clean_split_part(", ".join(parts))
    note = clean_split_part(", ".join(note_parts))
    return address, note


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_name(value: Any) -> str:
    name = clean_text(value)
    if not name or name.lower() in {"mijoz", "noma'lum", "nomalum", "unknown", "customer"}:
        return "MIJOZ"
    return name


def clean_address(value: Any, recipient_location: str) -> str:
    address = clean_text(value)
    if address:
        return address
    if recipient_location:
        return f"{recipient_location} markazi"
    return ""


def format_recipient_address(value: Any, recipient_location: str, delivery_type: str) -> tuple[str, str]:
    if delivery_type == "ДО ОФИСА":
        return branch_code_for_address(recipient_location, value)

    address = clean_text(value)
    if address:
        return address, ""
    if recipient_location:
        return f"{recipient_location} markazi", ""
    return "", ""


def resolve_allowed_recipient_location(customer: dict[str, Any]) -> tuple[str, str]:
    """P ustun uchun справочникdagi shahar/tuman nomini aniqlaydi.

    Manzil, izoh va AI/Excel dan kelgan hudud nomi ketma-ket tekshiriladi.
    """

    match = resolve_location(
        clean_text(customer.get("address")),
        clean_text(customer.get("note")),
        clean_text(customer.get("recipient_region_ru")),
    )
    return match.server, match.note


def parse_bool(value: str) -> str | None:
    normalized = value.strip().lower()
    true_values = {"true", "1", "ha", "xa", "yes", "y", "да"}
    false_values = {"false", "0", "yo'q", "yoq", "yuq", "no", "n", "нет"}
    if normalized in true_values:
        return "True"
    if normalized in false_values:
        return "False"
    return None


def chat_reply_lock(chat_id: int) -> asyncio.Lock:
    lock = reply_locks.get(chat_id)
    if lock is None:
        lock = asyncio.Lock()
        reply_locks[chat_id] = lock
    return lock


async def safe_answer(message: Message, text: str, **kwargs: Any) -> Any:
    async with chat_reply_lock(message.chat.id):
        while True:
            try:
                return await message.answer(text, **kwargs)
            except TelegramRetryAfter as error:
                await asyncio.sleep(error.retry_after + 1)


async def safe_send_message(bot: Bot, chat_id: int, text: str, **kwargs: Any) -> Any:
    async with chat_reply_lock(chat_id):
        while True:
            try:
                return await bot.send_message(chat_id, text, **kwargs)
            except TelegramRetryAfter as error:
                await asyncio.sleep(error.retry_after + 1)


async def safe_answer_document(message: Message, document: BufferedInputFile, **kwargs: Any) -> Any:
    async with chat_reply_lock(message.chat.id):
        while True:
            try:
                return await message.answer_document(document, **kwargs)
            except TelegramRetryAfter as error:
                await asyncio.sleep(error.retry_after + 1)


async def safe_edit_text(message: Message, text: str, **kwargs: Any) -> None:
    async with chat_reply_lock(message.chat.id):
        while True:
            try:
                await message.edit_text(text, **kwargs)
                return
            except TelegramRetryAfter as error:
                await asyncio.sleep(error.retry_after + 1)
            except Exception as error:
                logger.warning("Progress message edit failed: %s", error)
                return


def remember_setup_message(chat_id: int, message: Message | None) -> None:
    if message is None:
        return

    state = setup_states.get(chat_id)
    if state is None:
        return

    messages = state.setdefault("cleanup_messages", [])
    if any(saved.message_id == message.message_id for saved in messages):
        return
    messages.append(message)


async def cleanup_setup_messages(chat_id: int) -> None:
    state = setup_states.get(chat_id)
    if state is None:
        return

    messages = state.get("cleanup_messages", [])
    state["cleanup_messages"] = []
    for message in messages:
        while True:
            try:
                await message.delete()
                break
            except TelegramRetryAfter as error:
                await asyncio.sleep(error.retry_after + 1)
            except Exception as error:
                logger.debug("Setup message cleanup failed: %s", error)
                break


async def maybe_send_success_notice(message: Message, count: int) -> None:
    now = asyncio.get_running_loop().time()
    last_sent = last_success_notice_at.get(message.chat.id, 0)
    if now - last_sent < SUCCESS_NOTICE_INTERVAL_SECONDS:
        return

    last_success_notice_at[message.chat.id] = now
    await safe_answer(
        message,
        f"{count} ta mijoz Excel faylga qo'shildi.\n"
        "Ko'p xabar yuborilganda bot javoblarni kamaytiradi. Faylni olish uchun /excel yuboring.",
    )


def user_label_from_user(user: Any, fallback_id: int | None = None) -> str:
    if user is None:
        return f"id={fallback_id or 'nomaʼlum'}"
    name = " ".join(part for part in [user.first_name, user.last_name] if part)
    username = f"@{user.username}" if user.username else "username yo'q"
    return f"{name or 'Nomalum'} ({username}, id={user.id})"


def user_label(message: Message) -> str:
    return user_label_from_user(message.from_user, message.chat.id)


def access_phone_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📱 Telefon raqamni yuborish", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def remember_access_user(message: Message, phone: str = "") -> dict[str, Any]:
    user = message.from_user
    user_id = user.id if user else message.chat.id
    previous = access_requests.get(str(user_id), {})
    data = {
        "user_id": user_id,
        "first_name": clean_text(getattr(user, "first_name", "")),
        "last_name": clean_text(getattr(user, "last_name", "")),
        "username": clean_text(getattr(user, "username", "")),
        "phone": phone or clean_text(previous.get("phone")),
        "requested_at": datetime.now().isoformat(timespec="seconds"),
    }
    access_requests[str(user_id)] = data
    save_access_requests(access_requests)
    return data


def access_request_text(user_data: dict[str, Any]) -> str:
    full_name = " ".join(
        part
        for part in [clean_text(user_data.get("first_name")), clean_text(user_data.get("last_name"))]
        if part
    ) or "Noma'lum"
    username = f"@{user_data['username']}" if clean_text(user_data.get("username")) else "username yo'q"
    phone = clean_text(user_data.get("phone")) or "telefon yuborilmagan"
    return (
        "🔐 Botdan foydalanish uchun yangi so'rov\n\n"
        f"👤 Ism: {full_name}\n"
        f"🔗 Username: {username}\n"
        f"📞 Telefon: {phone}\n"
        f"🆔 User ID: {user_data.get('user_id')}"
    )


async def send_access_request_to_admins(message: Message, user_data: dict[str, Any]) -> None:
    user_id = int(user_data["user_id"])
    now = asyncio.get_running_loop().time()
    last_sent = access_request_sent_at.get(user_id, 0)
    if now - last_sent < ACCESS_REQUEST_INTERVAL_SECONDS:
        return
    access_request_sent_at[user_id] = now
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Ruxsat berish", callback_data=f"access:approve:{user_id}"),
                InlineKeyboardButton(text="⛔️ Rad etish", callback_data=f"access:deny:{user_id}"),
            ]
        ]
    )
    for admin_id in ADMIN_IDS:
        await safe_send_message(
            message.bot,
            admin_id,
            access_request_text(user_data),
            reply_markup=keyboard,
        )


async def request_access(message: Message) -> None:
    if not ADMIN_IDS:
        return

    user_id = message.from_user.id if message.from_user else message.chat.id
    user_data = remember_access_user(message)
    if clean_text(user_data.get("phone")):
        await send_access_request_to_admins(message, user_data)
        await safe_answer(
            message,
            "🔐 Ruxsat so'rovingiz adminga yuborildi.\n\n"
            "Admin tasdiqlagandan keyin /start ni qayta bosing.",
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    await safe_answer(
        message,
        "🔐 Botdan foydalanish uchun admin ruxsati kerak.\n\n"
        "So'rov yuborish uchun pastdagi tugma orqali telefon raqamingizni yuboring.",
        reply_markup=access_phone_keyboard(),
    )


async def ensure_user_access(message: Message) -> bool:
    user_id = message.from_user.id if message.from_user else message.chat.id
    if has_bot_access(user_id):
        return True
    await request_access(message)
    return False


async def ensure_callback_access(callback: CallbackQuery) -> bool:
    user_id = callback.from_user.id
    if has_bot_access(user_id):
        return True
    await callback.answer("Avval admin ruxsati kerak.", show_alert=True)
    if callback.message:
        await safe_send_message(
            callback.bot,
            callback.message.chat.id,
            "🔐 Botdan foydalanish uchun admin ruxsati kerak.\n\n"
            "So'rov yuborish uchun /start ni bosing va telefon raqamingizni yuboring.",
            reply_markup=access_phone_keyboard(),
        )
    return False


async def send_main_menu(message: Message, text: str | None = None) -> None:
    await safe_answer(
        message,
        text
        or (
            "Asosiy menyu.\n\n"
            "Excel ga yig'ish - yangi jo'natmalar ro'yxatini yig'ish.\n"
            "Ofislar ro'yxati - viloyat bo'yicha filiallarni ko'rish.\n"
            "Kalkulyator - EMU API orqali narx hisoblash.\n"
            "AI yordamchi - savol berib kerakli bo'limdan ma'lumot olish.\n"
            "Arxiv - tayyor Excel va shablon fayllar.\n"
            "Sozlamalar - jo'natuvchi ma'lumotlari va ruxsat holati."
        ),
        reply_markup=main_menu_keyboard(),
    )


async def access_callback_handler(callback: CallbackQuery) -> None:
    data = callback.data or ""
    parts = data.split(":", 2)
    admin_id = callback.from_user.id

    if len(parts) != 3 or parts[0] != "access":
        await callback.answer()
        return
    if not is_admin(admin_id):
        await callback.answer("Bu amal faqat admin uchun.", show_alert=True)
        return
    if not parts[2].isdigit():
        await callback.answer("User ID noto'g'ri.", show_alert=True)
        return

    action = parts[1]
    user_id = int(parts[2])
    user_data = access_requests.get(str(user_id), {"user_id": user_id})
    if action == "approve":
        approved_user_ids.add(user_id)
        save_approved_user_ids(approved_user_ids)
        await callback.answer("Ruxsat berildi.")
        if callback.message:
            await callback.message.edit_text(access_request_text(user_data) + "\n\n✅ Ruxsat berildi.")
        await safe_send_message(
            callback.bot,
            user_id,
            "Ruxsat berildi.\n\n"
            "Endi botdan foydalanishingiz mumkin. Boshlash uchun /start ni bosing.",
        )
        return

    if action == "deny":
        approved_user_ids.discard(user_id)
        save_approved_user_ids(approved_user_ids)
        await callback.answer("So'rov rad etildi.")
        if callback.message:
            await callback.message.edit_text(access_request_text(user_data) + "\n\n⛔️ So'rov rad etildi.")
        await safe_send_message(callback.bot, user_id, "Botdan foydalanish so'rovingiz rad etildi.")
        return

    await callback.answer("Noma'lum amal.", show_alert=True)


def normalize_cipher_prefix(value: str) -> str:
    prefix = re.sub(r"\s+", "", value.strip()).upper()
    return re.sub(r"[^A-ZА-ЯЁ0-9_-]", "", prefix)


def find_last_data_row(sheet: Any) -> int:
    for row_index in range(sheet.max_row, 1, -1):
        if any(sheet.cell(row_index, col).value not in (None, "") for col in range(1, sheet.max_column + 1)):
            return row_index
    return 1


def used_cipher_prefixes(sheet: Any) -> set[str]:
    prefixes: set[str] = set()
    for row_index in range(2, sheet.max_row + 1):
        code = clean_text(sheet.cell(row_index, 6).value)
        match = re.match(r"^(.+?)(\d+)$", code)
        if match:
            prefixes.add(match.group(1).upper())
    return prefixes


def next_cipher_index(sheet: Any, prefix: str) -> int:
    highest = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.IGNORECASE)
    for row_index in range(2, sheet.max_row + 1):
        code = clean_text(sheet.cell(row_index, 6).value)
        match = pattern.match(code)
        if match:
            highest = max(highest, int(match.group(1)))
    return highest + 1


def copy_row_style(sheet: Any, source_row: int, target_row: int, column_count: int) -> None:
    if source_row > sheet.max_row:
        source_row = 1
    for col in range(1, column_count + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def is_cipher_prefix_available(prefix: str) -> bool:
    existing_prefixes: set[str] = set()
    for client_type in (CLIENT_TYPE_PHYSICAL, CLIENT_TYPE_LEGAL):
        ensure_excel_file(client_type)
        workbook = load_workbook(excel_path_for_client_type(client_type))
        try:
            sheet = workbook.active
            existing_prefixes.update(used_cipher_prefixes(sheet))
        finally:
            workbook.close()
    active_prefixes = {
        session.get("cipher_prefix", "").upper()
        for session in sender_sessions.values()
        if session.get("cipher_prefix")
    }
    return prefix.upper() not in existing_prefixes and prefix.upper() not in active_prefixes


def validate_setup_value(chat_id: int, key: str, value: str) -> tuple[str | None, str | None]:
    value = clean_text(value)
    if not value and key in {"cipher_prefix"}:
        return "", None
    if not value:
        return None, "Bu maydon bo'sh bo'lmasin. Iltimos, qayta kiriting."

    if key in BUTTON_SETUP_OPTIONS:
        option_by_label = {
            option_label: option_value
            for option_label, option_value in BUTTON_SETUP_OPTIONS[key]
        }
        option_values = {option_value for _label, option_value in BUTTON_SETUP_OPTIONS[key]}
        if value in option_by_label:
            return option_by_label[value], None
        if value in option_values:
            return value, None
        if key == "cipher_prefix" and value:
            pass
        else:
            return None, "Iltimos, pastdagi tugmalardan birini tanlang."

    if key == "sender_phone":
        normalized, review = normalize_phone(value)
        if review:
            return None, "Telefon raqam noaniq. Masalan: 998901234567 yoki +998 90 123 45 67"
        return normalized, None

    if key == "cipher_prefix":
        if not value:
            return "", None
        prefix = normalize_cipher_prefix(value)
        if not prefix:
            return None, "Shifr faqat harf/raqamlardan iborat bo'lsin. Masalan: ABC"
        current_session = sender_sessions.get(chat_id, {})
        if current_session.get("cipher_prefix", "").upper() == prefix:
            return prefix, None
        if not is_cipher_prefix_available(prefix):
            return None, f"{prefix} shifri oldin ishlatilgan. Boshqa prefix kiriting."
        return prefix, None

    if key == "places_count":
        digits = re.sub(r"\D", "", value)
        if not digits or int(digits) < 1:
            return None, "Jo'natma soni 1 yoki undan katta raqam bo'lishi kerak."
        return str(int(digits)), None

    return value, None


def setup_step_keyboard(key: str) -> ReplyKeyboardMarkup | None:
    options = BUTTON_SETUP_OPTIONS.get(key)
    if not options:
        return reply_keyboard([], add_back=True)

    return reply_keyboard([label for label, _value in options], row_size=2)


async def ask_setup_step(message_or_query: Message | CallbackQuery, step_index: int) -> None:
    key, question = SETUP_STEPS[step_index]
    keyboard = setup_step_keyboard(key)
    sent_message = None
    chat_id = None

    if key == "cipher_prefix":
        question = f"{question}\n\nShifr kerak bo'lmasa, pastdagi tugmani bosing."

    if isinstance(message_or_query, CallbackQuery):
        if message_or_query.message:
            chat_id = message_or_query.message.chat.id
            sent_message = await safe_answer(message_or_query.message, question, reply_markup=keyboard)
    else:
        chat_id = message_or_query.chat.id
        sent_message = await safe_answer(message_or_query, question, reply_markup=keyboard)

    if chat_id is not None:
        remember_setup_message(chat_id, sent_message)


def setup_summary(session: dict[str, str]) -> str:
    client_type = "Yuridik mijoz" if session.get("client_type") == CLIENT_TYPE_LEGAL else "ФИЗ ЛИЦО"
    return (
        "Jo'natuvchi ma'lumotlari saqlandi:\n"
        f"Yo'nalish: {client_type}\n"
        f"Ism familiya: {session['sender_full_name']}\n"
        f"Telefon: {session['sender_phone']}\n"
        f"Manzil: {session['sender_address']}\n"
        f"Shahar: {session['sender_city_ru']}\n"
        f"Shifr: {session['cipher_prefix'] + '1, ' + session['cipher_prefix'] + '2, ...' if session['cipher_prefix'] else 'yoq'}\n"
        f"Yetkazib berish turi: {session['delivery_type']}\n"
        f"Оплата получателем: {session['payment_by_receiver']}\n"
        f"Og'irlik: {session['parcel_weight']}\n"
        f"Количество мест: {session['places_count']}\n\n"
        "Endi mijozlar ro'yxatini matn yoki rasm qilib yuboring."
    )


def legal_sender_defaults() -> dict[str, str]:
    return {
        "client_type": CLIENT_TYPE_LEGAL,
        "sender_full_name": "",
        "sender_phone": "",
        "sender_address": "",
        "sender_city_ru": "",
    }


async def start_setup(
    message: Message,
    reset: bool = False,
    client_type: str = CLIENT_TYPE_PHYSICAL,
    start_step: int = 0,
    initial_data: dict[str, str] | None = None,
) -> None:
    chat_id = message.chat.id
    if reset:
        sender_sessions.pop(chat_id, None)
    setup_states[chat_id] = {
        "step": start_step,
        "data": initial_data or {"client_type": client_type},
        "cleanup_messages": [],
    }
    intro_text = (
        "Yuridik mijoz tanlandi.\n"
        "Jo'natuvchi ma'lumotlari so'ralmaydi. Endi jo'natma sozlamalarini kiritamiz."
        if client_type == CLIENT_TYPE_LEGAL
        else "ФИЗ ЛИЦО tanlandi.\nExcel yaratishdan oldin jo'natuvchi ma'lumotlarini kiritamiz."
    )
    intro_message = await safe_answer(
        message,
        intro_text,
    )
    remember_setup_message(chat_id, intro_message)
    await ask_setup_step(message, start_step)


async def start_legal_setup(message: Message, reset: bool = True) -> None:
    await start_setup(
        message,
        reset=reset,
        client_type=CLIENT_TYPE_LEGAL,
        start_step=4,
        initial_data=legal_sender_defaults(),
    )


async def save_setup_value_and_advance(
    message_or_query: Message | CallbackQuery,
    chat_id: int,
    key: str,
    value: str,
) -> None:
    state = setup_states[chat_id]
    state["data"][key] = value
    step_index = state["step"] + 1

    if step_index >= len(SETUP_STEPS):
        sender_sessions[chat_id] = state["data"]
        setup_states.pop(chat_id, None)
        summary = setup_summary(sender_sessions[chat_id])
        if isinstance(message_or_query, CallbackQuery):
            if message_or_query.message:
                await safe_answer(message_or_query.message, summary, reply_markup=collect_active_keyboard())
        else:
            await safe_answer(message_or_query, summary, reply_markup=collect_active_keyboard())
        return

    state["step"] = step_index
    await ask_setup_step(message_or_query, step_index)


async def handle_setup_message(message: Message) -> bool:
    chat_id = message.chat.id
    state = setup_states.get(chat_id)
    if state is None:
        return False

    step_index = state["step"]
    key, _question = SETUP_STEPS[step_index]
    parsed_value, error = validate_setup_value(chat_id, key, message.text or "")
    if error:
        await safe_answer(message, error)
        if key in BUTTON_SETUP_OPTIONS:
            await ask_setup_step(message, step_index)
        return True

    remember_setup_message(chat_id, message)
    await cleanup_setup_messages(chat_id)
    await save_setup_value_and_advance(message, chat_id, key, parsed_value or "")
    return True


async def setup_callback_handler(callback: CallbackQuery) -> None:
    if not await ensure_callback_access(callback):
        return
    data = callback.data or ""
    parts = data.split(":", 2)
    if len(parts) != 3 or parts[0] != "setup":
        await callback.answer()
        return

    chat_id = callback.message.chat.id if callback.message else callback.from_user.id
    state = setup_states.get(chat_id)
    if state is None:
        await callback.answer("Sozlash jarayoni topilmadi. /setup ni bosing.", show_alert=True)
        return

    expected_key, _question = SETUP_STEPS[state["step"]]
    callback_key, value = parts[1], parts[2]
    if callback_key != expected_key:
        await callback.answer("Bu tugma eski savol uchun. Hozirgi savolga javob bering.", show_alert=True)
        return

    parsed_value, error = validate_setup_value(chat_id, expected_key, value)
    if error:
        await callback.answer(error, show_alert=True)
        return

    if callback.message:
        remember_setup_message(chat_id, callback.message)
    await callback.answer(f"Tanlandi: {parsed_value}")
    await cleanup_setup_messages(chat_id)
    await save_setup_value_and_advance(callback, chat_id, expected_key, parsed_value or "")


def prepare_rows(customers: list[dict[str, Any]], sender: dict[str, str]) -> list[list[str]]:
    rows = []
    is_legal = sender.get("client_type") == CLIENT_TYPE_LEGAL
    for customer in customers:
        recipient_location, location_review = resolve_allowed_recipient_location(customer)
        cleaned_address, address_note = split_address_and_note(customer.get("address"))
        note = "; ".join(
            part
            for part in [
                strip_phone_candidates(customer.get("note")),
                address_note,
            ]
            if part
        )
        normalized_phone, phone_review = normalize_phone_list(
            clean_text(customer.get("phone")),
            clean_text(customer.get("note")),
        )
        recipient_address, branch_code_review = format_recipient_address(
            cleaned_address,
            recipient_location,
            sender["delivery_type"],
        )
        review_parts = [
            clean_text(customer.get("needs_review")),
            phone_review,
            location_review,
            branch_code_review,
        ]
        review = "; ".join(part for part in review_parts if part)

        common_values = [
            "",
            clean_name(customer.get("full_name")),
            clean_name(customer.get("full_name")),
            recipient_address,
            normalized_phone,
            clean_text(customer.get("source_cipher")),
            sender["parcel_weight"],
            note,
            sender["places_count"],
            sender["delivery_type"],
        ]
        if is_legal:
            row = common_values + [
                recipient_location,
                sender["payment_by_receiver"],
            ]
        else:
            row = common_values + [
                sender["sender_full_name"],
                sender["sender_full_name"],
                sender["sender_address"],
                sender["sender_phone"],
                sender["sender_city_ru"],
                recipient_location,
                sender["payment_by_receiver"],
            ]
        row.append(review)
        rows.append(row)
    return rows


async def append_customers(customers: list[dict[str, Any]], sender: dict[str, str]) -> int:
    rows = prepare_rows(customers, sender)
    if not rows:
        return 0

    async with excel_lock:
        client_type = sender.get("client_type", CLIENT_TYPE_PHYSICAL)
        path = excel_path_for_client_type(client_type)
        headers = headers_for_client_type(client_type)
        ensure_excel_file(client_type)
        workbook = load_workbook(path)
        try:
            sheet = workbook.active

            next_row = find_last_data_row(sheet) + 1
            next_number = next_row - 1
            next_code_index = next_cipher_index(sheet, sender["cipher_prefix"]) if sender["cipher_prefix"] else 1
            for row in rows:
                copy_row_style(sheet, 2, next_row, len(headers))
                row[0] = next_number
                generated_cipher = False
                if row[5]:
                    row[5] = clean_text(row[5])
                else:
                    row[5] = f"{sender['cipher_prefix']}{next_code_index}" if sender["cipher_prefix"] else ""
                    generated_cipher = bool(row[5])
                review = row.pop()
                if review:
                    row[7] = "; ".join(part for part in [row[7], review] if part)
                for column_index, value in enumerate(row, start=1):
                    sheet.cell(next_row, column_index).value = value
                next_row += 1
                next_number += 1
                if generated_cipher:
                    next_code_index += 1

            workbook.save(path)
        finally:
            workbook.close()

    return len(rows)


async def get_excel_bytes(client_type: str = CLIENT_TYPE_PHYSICAL) -> bytes:
    async with excel_lock:
        ensure_excel_file(client_type)
        return excel_path_for_client_type(client_type).read_bytes()


def parse_openai_output(response: Any) -> list[dict[str, Any]]:
    output_text = getattr(response, "output_text", "")
    if not output_text:
        raise ValueError("OpenAI bo'sh javob qaytardi.")

    parsed = json.loads(output_text)
    customers = parsed.get("customers", [])
    if not isinstance(customers, list):
        raise ValueError("OpenAI javobida customers ro'yxati topilmadi.")

    return [item for item in customers if isinstance(item, dict)]


def call_openai_with_text(text: str) -> list[dict[str, Any]]:
    if openai_client is None:
        raise RuntimeError("OPENAI_API_KEY environment variable sozlanmagan.")

    response = openai_client.responses.create(
        model=OPENAI_MODEL,
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": f"Quyidagi mijoz ma'lumotlarini ajrating:\n\n{text}",
            },
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "customer_extraction",
                "schema": CUSTOMER_SCHEMA,
                "strict": True,
            }
        },
    )
    return parse_openai_output(response)


def call_openai_with_image(image_bytes: bytes, mime_type: str) -> list[dict[str, Any]]:
    if openai_client is None:
        raise RuntimeError("OPENAI_API_KEY environment variable sozlanmagan.")

    encoded = base64.b64encode(image_bytes).decode("utf-8")
    data_url = f"data:{mime_type};base64,{encoded}"

    response = openai_client.responses.create(
        model=OPENAI_MODEL,
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": "Rasmdagi mijoz ma'lumotlarini maksimal aniqlik bilan o'qing va ajrating.",
                    },
                    {
                        "type": "input_image",
                        "image_url": data_url,
                    },
                ],
            },
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "customer_extraction",
                "schema": CUSTOMER_SCHEMA,
                "strict": True,
            }
        },
    )
    return parse_openai_output(response)


def is_excel_document(file_name: str, mime_type: str) -> bool:
    lowered_name = clean_text(file_name).casefold()
    lowered_mime = clean_text(mime_type).casefold()
    return lowered_name.endswith((".xlsx", ".xlsm")) or lowered_mime in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }


def looks_like_excel_header(values: list[str]) -> bool:
    normalized = [normalize_location_key(value) for value in values]
    joined = " ".join(normalized)
    header_words = {"shifr", "cipher", "telefon", "phone", "tel", "ism", "fio", "address", "manzil"}
    return sum(1 for word in header_words if word in joined) >= 2


def extract_customers_from_excel_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    customers: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [clean_text(value) for value in row]
                while values and not values[-1]:
                    values.pop()
                if not any(values):
                    continue
                if row_index == 1 and looks_like_excel_header(values):
                    continue

                cipher = values[0] if len(values) > 0 else ""
                phone = values[1] if len(values) > 1 else ""
                full_name = values[2] if len(values) > 2 else ""
                address = values[3] if len(values) > 3 else ""
                extra_values = [value for value in values[4:] if value]
                if cipher and not any([phone, full_name, address]):
                    continue
                if len(values) < 4:
                    extra_values.append("Excel qatorida 4 ta asosiy ustun to'liq emas")

                if not any([cipher, phone, full_name, address]):
                    continue

                location = resolve_server_location(address)
                customers.append(
                    {
                        "number": "",
                        "source_cipher": cipher,
                        "full_name": full_name,
                        "phone": phone,
                        "address": address,
                        "recipient_region_ru": location,
                        "note": "; ".join(extra_values),
                        "needs_review": "" if address and phone else f"Excel {sheet.title}!{row_index}-qatorni tekshirish kerak",
                    }
                )
    finally:
        workbook.close()

    if not customers:
        raise ValueError("Excel faylda mijoz qatorlari topilmadi.")
    return customers


async def handle_customers(message: Message, customers: list[dict[str, Any]]) -> None:
    sender = sender_sessions.get(message.chat.id)
    if sender is None:
        await safe_answer(message, "Avval jo'natuvchi ma'lumotlarini kiritish kerak.")
        await start_setup(message)
        return

    if not customers:
        await safe_answer(
            message,
            "Mijoz ma'lumotlari topilmadi. Iltimos, matnni aniqroq yuboring yoki rasm sifatini yaxshilang."
        )
        return

    count = await append_customers(customers, sender)
    await maybe_send_success_notice(message, count)


async def enqueue_batch_item(item: BatchItem) -> None:
    chat_id = item.message.chat.id
    state = batch_states.get(chat_id)
    if state is None:
        state = BatchState()
        batch_states[chat_id] = state

    state.items.append(item)
    state.last_added_at = asyncio.get_running_loop().time()

    if state.task is None or state.task.done():
        state.task = asyncio.create_task(process_batch(chat_id))


async def wait_for_batch_idle(chat_id: int) -> BatchState | None:
    while True:
        state = batch_states.get(chat_id)
        if state is None:
            return None

        elapsed = asyncio.get_running_loop().time() - state.last_added_at
        if elapsed >= BATCH_IDLE_SECONDS:
            return state

        await asyncio.sleep(max(0.2, BATCH_IDLE_SECONDS - elapsed))


async def extract_customers_from_batch_item(item: BatchItem) -> list[dict[str, Any]]:
    if item.kind == "text":
        return await asyncio.to_thread(call_openai_with_text, item.text)

    if item.bot is None:
        raise RuntimeError("Bot instance topilmadi.")

    file = await item.bot.get_file(item.file_id)
    buffer = io.BytesIO()
    await item.bot.download_file(file.file_path, destination=buffer)
    file_bytes = buffer.getvalue()

    if item.kind == "excel":
        return await asyncio.to_thread(extract_customers_from_excel_bytes, file_bytes)

    return await asyncio.to_thread(
        call_openai_with_image,
        file_bytes,
        item.mime_type or "image/jpeg",
    )


def batch_progress_text(total: int, processed: int, added: int, errors: list[str]) -> str:
    return (
        f"{total} ta ma'lumot qabul qilindi\n"
        f"{processed}/{total} tahlil qilindi\n"
        f"Excelga qo'shilgan mijozlar: {added}\n"
        f"Xatoliklar: {len(errors)}"
    )


def batch_final_text(total: int, added: int, errors: list[str]) -> str:
    lines = [
        "Tahlil tugadi.",
        f"Qabul qilingan xabarlar: {total}",
        f"Excelga qo'shilgan mijozlar: {added}",
    ]
    if errors:
        lines.append(f"Xatoliklar: {len(errors)}")
        lines.extend(errors[:10])
        if len(errors) > 10:
            lines.append(f"... yana {len(errors) - 10} ta xatolik bor")
    else:
        lines.append("Xatoliklar: yo'q")
    lines.append("Excel faylni olish uchun /excel yuboring.")
    return "\n".join(lines)


async def process_batch(chat_id: int) -> None:
    state = await wait_for_batch_idle(chat_id)
    if state is None or not state.items:
        batch_states.pop(chat_id, None)
        return

    items = list(state.items)
    state.items.clear()

    first_message = items[0].message
    progress_message = await safe_answer(
        first_message,
        batch_progress_text(len(items), 0, 0, []),
    )

    sender = sender_sessions.get(chat_id)
    errors: list[str] = []
    processed_total = 0
    added_total = 0
    extracted_by_index: list[list[dict[str, Any]]] = [[] for _ in items]
    last_progress_edit_at = 0.0

    if sender is None:
        await safe_edit_text(
            progress_message,
            batch_final_text(
                len(items),
                0,
                ["Jo'natuvchi ma'lumotlari sozlanmagan. /setup ni bosing."],
            ),
        )
        batch_states.pop(chat_id, None)
        return

    semaphore = asyncio.Semaphore(BATCH_CONCURRENCY)

    async def extract_with_index(index: int, item: BatchItem) -> tuple[int, list[dict[str, Any]] | None, str | None]:
        async with semaphore:
            try:
                customers = await extract_customers_from_batch_item(item)
                if not customers:
                    raise RuntimeError("mijoz ma'lumotlari topilmadi")
                return index, customers, None
            except Exception as error:
                logger.exception("Batch item failed")
                return index, None, str(error)

    tasks = [
        asyncio.create_task(extract_with_index(index, item))
        for index, item in enumerate(items, start=1)
    ]

    for task in asyncio.as_completed(tasks):
        index, customers, error = await task
        processed_total += 1
        if customers is not None:
            extracted_by_index[index - 1] = customers
        if error:
            errors.append(f"{index}-xabar: {error}")

        now = asyncio.get_running_loop().time()
        if processed_total == len(items) or now - last_progress_edit_at >= BATCH_PROGRESS_EDIT_INTERVAL_SECONDS:
            last_progress_edit_at = now
            await safe_edit_text(
                progress_message,
                batch_progress_text(len(items), processed_total, added_total, errors),
            )

    customers_to_append = [
        customer
        for customers in extracted_by_index
        for customer in customers
    ]
    if customers_to_append:
        try:
            added_total = await append_customers(customers_to_append, sender)
        except Exception as error:
            logger.exception("Batch append failed")
            errors.append(f"Excelga yozishda xatolik: {error}")

    await safe_edit_text(
        progress_message,
        batch_final_text(len(items), added_total, errors),
    )

    state = batch_states.get(chat_id)
    if state and state.items:
        state.task = asyncio.create_task(process_batch(chat_id))
    else:
        batch_states.pop(chat_id, None)


async def start_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await send_main_menu(
        message,
        "Assalomu alaykum!\n\n"
        "Bot mijoz ma'lumotlarini matn, rasm yoki Excel fayldan ajratib, Excel shablonga yozadi.\n"
        "Quyidagi bo'limlardan birini tanlang.",
    )


async def access_contact_handler(message: Message) -> None:
    user_id = message.from_user.id if message.from_user else message.chat.id
    if has_bot_access(user_id):
        await send_main_menu(message, "Sizda ruxsat bor. Asosiy menyu:")
        return

    contact = message.contact
    if contact is None:
        await request_access(message)
        return
    if contact.user_id and contact.user_id != user_id:
        await safe_answer(
            message,
            "Iltimos, o'zingizning telefon raqamingizni yuboring.",
            reply_markup=access_phone_keyboard(),
        )
        return

    normalized_phone, review = normalize_phone(clean_text(contact.phone_number))
    phone = normalized_phone if not review else clean_text(contact.phone_number)
    user_data = remember_access_user(message, phone)
    await send_access_request_to_admins(message, user_data)
    await safe_answer(
        message,
        "✅ Telefon raqamingiz qabul qilindi.\n\n"
        "🔐 Ruxsat so'rovi adminga yuborildi. Tasdiqlangandan keyin /start ni bosing.",
        reply_markup=ReplyKeyboardRemove(),
    )


async def help_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await safe_answer(
        message,
        "Foydalanish yo'riqnomasi:\n\n"
        "1. Excel ga yig'ish bo'limiga kiring.\n"
        "2. Yuridik mijoz yoki ФИЗ ЛИЦО yo'nalishini tanlang.\n"
        "3. Bot so'ragan sozlamalarga javob bering.\n"
        "4. Mijoz ma'lumotlarini matn, rasm yoki .xlsx Excel qilib yuboring.\n"
        "5. Tayyor faylni Arxiv bo'limidan oling.\n\n"
        "Har bir ichki bo'limda Orqaga tugmasi bor.",
        reply_markup=main_menu_keyboard(),
    )


async def setup_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await start_setup(message, reset=True)


async def offices_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await show_offices_menu(message)


async def calculator_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await show_calculator_menu(message)


async def ai_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await show_ai_assistant(message)


async def excel_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    client_type = current_client_type(message.chat.id)
    file_bytes = await get_excel_bytes(client_type)
    filename = excel_path_for_client_type(client_type).name
    await safe_answer_document(
        message,
        BufferedInputFile(file_bytes, filename=filename),
        caption="Yangilangan mijozlar ro'yxati.",
    )


async def template_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    if not TEMPLATE_PATH.exists():
        await safe_answer(message, "Shablon fayl topilmadi.")
        return

    await safe_answer_document(
        message,
        BufferedInputFile(TEMPLATE_PATH.read_bytes(), filename="yangi_shablon.xlsx"),
        caption="Excel shablon fayli.",
    )


async def current_templates_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return

    template_files = sorted(TEMPLATE_DIR.glob("*.xlsx"))
    if not template_files:
        await safe_answer(message, "Joriy shablon fayllari topilmadi.", reply_markup=settings_menu_keyboard())
        return

    await safe_answer(
        message,
        f"📑 Joriy shablonlar yuborilmoqda: {len(template_files)} ta fayl.",
        reply_markup=settings_menu_keyboard(),
    )
    for path in template_files:
        await safe_answer_document(
            message,
            BufferedInputFile(path.read_bytes(), filename=path.name),
            caption=f"📄 {path.name}",
        )


def format_emu_database_status() -> str:
    database = load_emu_database()
    updated_at = float(database.get("updated_at") or 0)
    if updated_at:
        updated_text = datetime.fromtimestamp(updated_at).strftime("%Y-%m-%d %H:%M")
        age_days = emu_database_age_seconds() / 86400
        age_text = f"{age_days:.1f} kun oldin"
    else:
        updated_text = "hali yangilanmagan"
        age_text = "noma'lum"

    calculator_cache = database.get("calculator_cache")
    calculator_count = len(calculator_cache) if isinstance(calculator_cache, dict) else 0
    return (
        "📚 EMU baza holati\n\n"
        f"Viloyatlar: {len(database.get('regions') or [])}\n"
        f"Shahar/tumanlar: {len(database.get('cities') or [])}\n"
        f"Ofislar: {len(database.get('branches') or [])}\n"
        f"Kalkulyator cache: {calculator_count}\n"
        f"Oxirgi yangilanish: {updated_text}\n"
        f"Yoshi: {age_text}"
    )


async def refresh_emu_database_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    user_id = message.from_user.id if message.from_user else message.chat.id
    if not is_admin(user_id):
        await safe_answer(message, "Bu amal faqat admin uchun.", reply_markup=settings_menu_keyboard())
        return

    status_message = await safe_answer(
        message,
        "🔄 EMU bazasi yangilanmoqda...\n\n"
        "Bu jarayonda bot emu.uz API'dan viloyat, tuman/shahar va ofislar ro'yxatini qayta oladi.",
        reply_markup=settings_menu_keyboard(),
    )
    try:
        database = await refresh_emu_database()
        await safe_edit_text(
            status_message,
            "✅ EMU bazasi yangilandi.\n\n"
            f"Viloyatlar: {len(database.get('regions') or [])}\n"
            f"Shahar/tumanlar: {len(database.get('cities') or [])}\n"
            f"Ofislar: {len(database.get('branches') or [])}",
        )
    except Exception as error:
        logger.exception("EMU database refresh failed")
        await safe_edit_text(status_message, f"⚠️ EMU bazani yangilashda xatolik: {error}")


async def show_offices_menu(message: Message) -> None:
    state = {"mode": "offices", "step": "region"}
    service_states[message.chat.id] = state
    try:
        regions = await get_emu_regions()
    except Exception as error:
        logger.exception("EMU regions loading failed")
        await safe_answer(message, f"Ofislar ro'yxatini olishda xatolik: {error}")
        return

    await safe_answer(
        message,
        "🏢 Ofislar ro'yxati\n\n"
        "📍 Viloyatni tanlang. Keyin shu hududdagi ofislar sahifalab chiqadi.",
        reply_markup=region_reply_keyboard(regions, state),
    )


async def show_calculator_menu(message: Message) -> None:
    state = {"mode": "calculator", "step": "sender_region"}
    service_states[message.chat.id] = state
    try:
        regions = await get_emu_regions()
    except Exception as error:
        logger.exception("EMU regions loading failed")
        await safe_answer(message, f"Kalkulyator ma'lumotlarini olishda xatolik: {error}")
        return

    await safe_answer(
        message,
        "🧮 Kalkulyator\n\n"
        "📦 Jo'natilish nuqtasining viloyatini tanlang.",
        reply_markup=region_reply_keyboard(regions, state),
    )


async def show_ai_assistant(message: Message) -> None:
    service_states[message.chat.id] = {"mode": "ai"}
    await safe_answer(
        message,
        "AI yordamchi bo'limi.\n\n"
        "EMU bo'yicha savolingizni yozing. Masalan:\n"
        "- Samarqand ofislari qayerda?\n"
        "- Andijondan Toshkentga 2 kg qancha?\n"
        "- Do ofisa va Na dom farqi nima?\n\n"
        f"Chiqish uchun {MENU_BACK} tugmasini bosing.",
        reply_markup=reply_keyboard([], add_back=True),
    )


async def emu_callback_handler(callback: CallbackQuery) -> None:
    if not await ensure_callback_access(callback):
        return
    data = callback.data or ""
    parts = data.split(":")
    chat_id = callback.message.chat.id if callback.message else callback.from_user.id

    if data == "emu:back":
        service_states.pop(chat_id, None)
        await callback.answer()
        if callback.message:
            await safe_edit_text(callback.message, "Asosiy menyuga qaytdingiz.")
            await safe_send_message(callback.bot, chat_id, "🏠 Asosiy menyu", reply_markup=main_menu_keyboard())
        return

    try:
        if len(parts) == 3 and parts[1] == "office_region":
            region_id = int(parts[2])
            branches = await get_emu_branches(region_id=region_id)
            regions = await get_emu_regions()
            region = next((item for item in regions if int(item.get("id") or 0) == region_id), {})
            title = f"{localized_name(region)} ofislari"
            state = service_states.setdefault(chat_id, {"mode": "offices"})
            state.update({"step": "branches", "branches": branches, "page": 0, "title": title})
            await callback.answer("Ofislar yuklandi.")
            if callback.message:
                await safe_edit_text(callback.message, "🏢 Ofislar pastki menyuga chiqarildi.")
                await safe_send_message(
                    callback.bot,
                    chat_id,
                    format_branches_page(branches, title, 0),
                    reply_markup=offices_page_keyboard(0, len(branches)),
                )
            return

        if len(parts) == 3 and parts[1] == "calc_sender_region":
            region_id = int(parts[2])
            state = service_states.setdefault(chat_id, {"mode": "calculator"})
            state.update({"sender_region_id": region_id})
            await callback.answer()
            if region_id == TASHKENT_REGION_ID:
                state.update({"sender_city_id": TASHKENT_CITY_ID, "step": "receiver_region"})
                regions = await get_emu_regions()
                if callback.message:
                    await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                    await safe_send_message(
                        callback.bot,
                        chat_id,
                        "Jo'natilish nuqtasi: Toshkent.\n\nEndi olish nuqtasining viloyatini tanlang.",
                        reply_markup=region_reply_keyboard(regions, state),
                    )
                return
            cities = await get_emu_cities(region_id)
            state["step"] = "sender_city"
            if callback.message:
                await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                await safe_send_message(
                    callback.bot,
                    chat_id,
                    "Jo'natilish nuqtasining tuman/shahrini tanlang.",
                    reply_markup=city_reply_keyboard(cities, state),
                )
            return

        if len(parts) == 3 and parts[1] == "calc_sender_city":
            city_id = int(parts[2])
            state = service_states.setdefault(chat_id, {"mode": "calculator"})
            state.update({"sender_city_id": city_id, "step": "receiver_region"})
            regions = await get_emu_regions()
            await callback.answer()
            if callback.message:
                await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                await safe_send_message(
                    callback.bot,
                    chat_id,
                    "Endi olish nuqtasining viloyatini tanlang.",
                    reply_markup=region_reply_keyboard(regions, state),
                )
            return

        if len(parts) == 3 and parts[1] == "calc_receiver_region":
            region_id = int(parts[2])
            state = service_states.setdefault(chat_id, {"mode": "calculator"})
            state.update({"receiver_region_id": region_id})
            await callback.answer()
            if region_id == TASHKENT_REGION_ID:
                state.update({"receiver_city_id": TASHKENT_CITY_ID, "step": "service"})
                if callback.message:
                    await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                    await safe_send_message(
                        callback.bot,
                        chat_id,
                        "Olish nuqtasi: Toshkent.\n\nYetkazib berish turini tanlang.",
                        reply_markup=calculator_service_reply_keyboard(),
                    )
                return
            cities = await get_emu_cities(region_id)
            state["step"] = "receiver_city"
            if callback.message:
                await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                await safe_send_message(
                    callback.bot,
                    chat_id,
                    "Olish nuqtasining tuman/shahrini tanlang.",
                    reply_markup=city_reply_keyboard(cities, state),
                )
            return

        if len(parts) == 3 and parts[1] == "calc_receiver_city":
            city_id = int(parts[2])
            state = service_states.setdefault(chat_id, {"mode": "calculator"})
            state.update({"receiver_city_id": city_id, "step": "service"})
            await callback.answer()
            if callback.message:
                await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                await safe_send_message(
                    callback.bot,
                    chat_id,
                    "Olish turi: ofisgachami yoki uygachami?",
                    reply_markup=calculator_service_reply_keyboard(),
                )
            return

        if len(parts) == 3 and parts[1] == "calc_service":
            service_id = int(parts[2])
            state = service_states.setdefault(chat_id, {"mode": "calculator"})
            state.update({"service_id": service_id, "step": "weight"})
            await callback.answer()
            if callback.message:
                await safe_edit_text(callback.message, "🧮 Tanlov pastki menyuga o'tkazildi.")
                await safe_send_message(
                    callback.bot,
                    chat_id,
                    "Jo'natmaning og'irligini kiriting.\n\n"
                    "Agar gabaritda o'lchangan og'irligi kattaroq bo'lsa, shuni kiriting. Masalan: 1.5",
                    reply_markup=reply_keyboard([], add_back=True),
                )
            return
    except Exception as error:
        logger.exception("EMU callback failed")
        await callback.answer("Xatolik yuz berdi.", show_alert=True)
        if callback.message:
            await safe_answer(callback.message, f"Xatolik: {error}")
        return

    await callback.answer()


async def show_collect_menu(message: Message) -> None:
    await safe_answer(
        message,
        "Jo'natmalarni yig'ish bo'limi.\n\n"
        "Yuridik mijoz - jo'natuvchi ma'lumotlari so'ralmaydi.\n"
        "ФИЗ ЛИЦО - jo'natuvchi ma'lumotlari odatdagidek so'raladi.\n\n"
        "Kerakli yo'nalishni tanlang.",
        reply_markup=collect_menu_keyboard(),
    )


async def show_archive_menu(message: Message) -> None:
    await safe_answer(
        message,
        "Arxiv bo'limi.\n\n"
        "Excel fayl - yig'ilgan mijozlar ro'yxatini yuboradi.\n"
        "Shablon - hozirgi Excel shablonni yuboradi.\n"
        "Ro'yxatni tozalash - Excel ro'yxatini boshidan boshlaydi.",
        reply_markup=archive_menu_keyboard(),
    )


async def show_settings_menu(message: Message) -> None:
    await safe_answer(
        message,
        "Sozlamalar bo'limi.\n\n"
        "Jo'natuvchi sozlamalari - ФИЗ ЛИЦО uchun ma'lumotlarni qayta kiritish.\n"
        "Ruxsat holati - botdan foydalanish ruxsatini ko'rsatadi.\n"
        "Joriy shablonlar - hozir ishlatilayotgan Excel shablon fayllarini yuboradi.\n"
        "EMU bazani yangilash - admin uchun, emu.uz'dan yangi filial va shahar ma'lumotlarini oladi.",
        reply_markup=settings_menu_keyboard(),
    )


async def handle_menu_message(message: Message) -> bool:
    text = MENU_ALIASES.get((message.text or "").strip(), (message.text or "").strip())
    if not text:
        return False

    if text == MENU_BACK:
        setup_states.pop(message.chat.id, None)
        service_states.pop(message.chat.id, None)
        await send_main_menu(message)
        return True

    if text == MENU_COLLECT:
        await show_collect_menu(message)
        return True

    if text == MENU_OFFICES:
        await show_offices_menu(message)
        return True

    if text == MENU_CALCULATOR:
        await show_calculator_menu(message)
        return True

    if text == MENU_AI_ASSISTANT:
        await show_ai_assistant(message)
        return True

    if text == MENU_ARCHIVE:
        await show_archive_menu(message)
        return True

    if text == MENU_SETTINGS:
        await show_settings_menu(message)
        return True

    if text == MENU_LEGAL:
        await start_legal_setup(message, reset=True)
        return True

    if text == MENU_PHYSICAL:
        await start_setup(message, reset=True, client_type=CLIENT_TYPE_PHYSICAL)
        return True

    if text == MENU_EXCEL_FILE:
        await excel_handler(message)
        return True

    if text == MENU_TEMPLATE_FILE:
        await template_handler(message)
        return True

    if text == MENU_CURRENT_TEMPLATES:
        await current_templates_handler(message)
        return True

    if text == MENU_CLEAR:
        await clear_handler(message, start_next_setup=False)
        return True

    if text == MENU_RESET_SETUP:
        await start_setup(message, reset=True, client_type=CLIENT_TYPE_PHYSICAL)
        return True

    if text == MENU_ACCESS_STATUS:
        user_id = message.from_user.id if message.from_user else message.chat.id
        status = "admin" if is_admin(user_id) else "ruxsat berilgan"
        if not ADMIN_IDS:
            status = "ruxsat tekshiruvi o'chirilgan"
        await safe_answer(message, f"Ruxsat holati: {status}", reply_markup=settings_menu_keyboard())
        return True

    if text == MENU_EMU_DB_STATUS:
        await safe_answer(message, format_emu_database_status(), reply_markup=settings_menu_keyboard())
        return True

    if text == MENU_REFRESH_EMU_DB:
        await refresh_emu_database_handler(message)
        return True

    return False


async def answer_ai_question(message: Message, question: str) -> None:
    lowered = question.lower()
    context_parts: list[str] = []

    try:
        if any(word in lowered for word in ["narx", "kalk", "qancha", "kg", "кг", "sum", "so'm", "som"]):
            calculator_answer = await calculate_from_ai_question(message, question)
            if calculator_answer:
                await safe_answer(message, calculator_answer)
                return

        if any(word in lowered for word in ["ofis", "filial", "office", "branch"]):
            branches = await get_all_emu_branches()
            matching = find_matching_branches_for_question(branches, question)
            if not matching:
                matching = branches
            if any(word in lowered for word in ["nechta", "qancha", "soni"]):
                await safe_answer(message, format_office_count_answer(matching))
                return
            if any(word in lowered for word in ["bormi", "qayerda", "manzil", "telefon", "ish vaqti"]):
                await safe_answer(message, format_branches_list(matching, "Ofislar bo'yicha topilgan ma'lumot", limit=10))
                return
            context_parts.append(format_branches_list(matching, "Ofislar bo'yicha topilgan ma'lumot", limit=20))

        if any(word in lowered for word in ["viloyat", "tuman", "shahar", "city", "region"]):
            cities = await get_emu_cities()
            sample = [
                f"{city.get('id')}: {localized_name(city)} ({clean_text((city.get('region_name') or {}).get('UZ'))})"
                for city in cities[:60]
            ]
            context_parts.append("EMU shahar/tuman ma'lumotlari:\n" + "\n".join(sample))

        if any(word in lowered for word in ["narx", "kalk", "qancha", "kg", "sum", "so'm", "som"]):
            context_parts.append(
                "Kalkulyator uchun aniq hisoblashda jo'natuvchi shahar/tuman, oluvchi shahar/tuman, "
                "yetkazish turi va og'irlik kerak. Botdagi Kalkulyator bo'limi shu ma'lumotlar asosida EMU API'dan narx oladi."
            )

        if not context_parts:
            context_parts.append(
                "EMU Express saytida ofislar, shahar/tumanlar, filial telefonlari, kalkulyator narxlari, "
                "Do ofisa/Na dom tariflari va indeks ma'lumotlari mavjud."
            )

        context = "\n\n".join(context_parts)[:8000]
        if openai_client is None:
            await safe_answer(message, context[:3500])
            return

        response = await asyncio.to_thread(
            openai_client.responses.create,
            model=OPENAI_MODEL,
            input=[
                {
                    "role": "system",
                    "content": (
                        "Siz EMU botining yordamchisisiz. Faqat berilgan kontekstga tayanib, "
                        "qisqa, amaliy va o'zbek tilida javob bering. Javobni abzastlarga ajrating: "
                        "asosiy xulosa, keyin kerak bo'lsa punktli ro'yxat. Har bir fakt alohida qatorda bo'lsin. "
                        "Agar ma'lumot yetmasa, qaysi bo'limdan foydalanish kerakligini ayting."
                    ),
                },
                {
                    "role": "user",
                    "content": f"Kontekst:\n{context}\n\nSavol:\n{question}",
                },
            ],
        )
        answer = clean_text(getattr(response, "output_text", ""))
        await safe_answer(message, answer or "Javob topilmadi. Savolni aniqroq yozing.")
    except Exception as error:
        logger.exception("AI assistant failed")
        await safe_answer(message, f"AI yordamchida xatolik: {error}")


async def send_offices_page(message: Message, state: dict[str, Any], page: int = 0) -> None:
    branches = state.get("filtered_branches") or state.get("branches") or []
    title = state.get("title") or "Ofislar"
    page_count = max(1, (len(branches) + OFFICES_PAGE_SIZE - 1) // OFFICES_PAGE_SIZE)
    page = max(0, min(page, page_count - 1))
    state["page"] = page
    await safe_answer(
        message,
        format_branches_page(branches, title, page),
        reply_markup=offices_page_keyboard(page, len(branches)),
    )


async def handle_service_text(message: Message) -> bool:
    state = service_states.get(message.chat.id)
    if not state:
        return False

    mode = state.get("mode")
    text = (message.text or "").strip()

    if text == MENU_CANCEL:
        service_states.pop(message.chat.id, None)
        await send_main_menu(message, "❌ Amal bekor qilindi.")
        return True

    if mode == "offices":
        step = state.get("step")
        if step == "region":
            region_id = selected_option_id(state, text)
            if region_id is None:
                await safe_answer(message, "📍 Pastdagi tugmalardan viloyatni tanlang.")
                return True
            try:
                branches = await get_emu_branches(region_id=region_id)
                regions = await get_emu_regions()
                region = next((item for item in regions if int(item.get("id") or 0) == region_id), {})
                state.update(
                    {
                        "step": "browser",
                        "region_id": region_id,
                        "branches": branches,
                        "filtered_branches": branches,
                        "title": f"{localized_name(region)} ofislari",
                        "page": 0,
                    }
                )
                await send_offices_page(message, state, 0)
            except Exception as error:
                logger.exception("Office list failed")
                await safe_answer(message, f"⚠️ Ofislar ro'yxatini olishda xatolik: {error}")
            return True

        if step == "browser":
            if text == MENU_NEXT_PAGE:
                await send_offices_page(message, state, int(state.get("page") or 0) + 1)
                return True
            if text == MENU_PREV_PAGE:
                await send_offices_page(message, state, int(state.get("page") or 0) - 1)
                return True
            if text == MENU_SEARCH:
                state["step"] = "search"
                await safe_answer(
                    message,
                    "🔎 Qidirish uchun tuman, filial nomi yoki manzil bo'lagini yozing.\n"
                    "Masalan: Sergeli, Olmazor, Qoraqamish",
                    reply_markup=reply_keyboard([MENU_CANCEL], row_size=1),
                )
                return True
            await safe_answer(message, "Pastdagi tugmalardan amal tanlang.")
            return True

        if step == "search":
            filtered = filter_branches(state.get("branches") or [], text)
            state["filtered_branches"] = filtered
            state["title"] = f"{state.get('title', 'Ofislar')} | qidiruv: {text}"
            state["step"] = "browser"
            await send_offices_page(message, state, 0)
            return True

    if mode == "ai":
        await answer_ai_question(message, text)
        return True

    if mode == "calculator" and state.get("step") != "weight":
        step = state.get("step")
        try:
            if step == "sender_region":
                region_id = selected_option_id(state, text)
                if region_id is None:
                    await safe_answer(message, "📍 Jo'natilish viloyatini pastdagi tugmalardan tanlang.")
                    return True
                state["sender_region_id"] = region_id
                if region_id == TASHKENT_REGION_ID:
                    state.update({"sender_city_id": TASHKENT_CITY_ID, "step": "receiver_region"})
                    regions = await get_emu_regions()
                    await safe_answer(
                        message,
                        "✅ Jo'natilish nuqtasi: Toshkent.\n\n📍 Endi olish nuqtasining viloyatini tanlang.",
                        reply_markup=region_reply_keyboard(regions, state),
                    )
                    return True
                cities = await get_emu_cities(region_id)
                state["step"] = "sender_city"
                await safe_answer(
                    message,
                    "🏙 Jo'natilish nuqtasining tuman/shahrini tanlang.",
                    reply_markup=city_reply_keyboard(cities, state),
                )
                return True

            if step == "sender_city":
                city_id = selected_option_id(state, text)
                if city_id is None:
                    await safe_answer(message, "🏙 Tuman/shaharni pastdagi tugmalardan tanlang.")
                    return True
                state.update({"sender_city_id": city_id, "step": "receiver_region"})
                regions = await get_emu_regions()
                await safe_answer(
                    message,
                    "📍 Endi olish nuqtasining viloyatini tanlang.",
                    reply_markup=region_reply_keyboard(regions, state),
                )
                return True

            if step == "receiver_region":
                region_id = selected_option_id(state, text)
                if region_id is None:
                    await safe_answer(message, "📍 Olish viloyatini pastdagi tugmalardan tanlang.")
                    return True
                state["receiver_region_id"] = region_id
                if region_id == TASHKENT_REGION_ID:
                    state.update({"receiver_city_id": TASHKENT_CITY_ID, "step": "service"})
                    await safe_answer(
                        message,
                        "✅ Olish nuqtasi: Toshkent.\n\n🚚 Yetkazib berish turini tanlang.",
                        reply_markup=calculator_service_reply_keyboard(),
                    )
                    return True
                cities = await get_emu_cities(region_id)
                state["step"] = "receiver_city"
                await safe_answer(
                    message,
                    "🏙 Olish nuqtasining tuman/shahrini tanlang.",
                    reply_markup=city_reply_keyboard(cities, state),
                )
                return True

            if step == "receiver_city":
                city_id = selected_option_id(state, text)
                if city_id is None:
                    await safe_answer(message, "🏙 Olish tuman/shahrini pastdagi tugmalardan tanlang.")
                    return True
                state.update({"receiver_city_id": city_id, "step": "service"})
                await safe_answer(
                    message,
                    "🚚 Olish turi: ofisgachami yoki uygachami?",
                    reply_markup=calculator_service_reply_keyboard(),
                )
                return True

            if step == "service":
                if text == MENU_DO_OFFICE or "ДО ОФИСА" in text:
                    service_id = 1
                elif text == MENU_TO_HOME or "НА ДОМ" in text:
                    service_id = 3
                else:
                    await safe_answer(message, "🚚 Yetkazib berish turini pastdagi tugmalardan tanlang.")
                    return True
                state.update({"service_id": service_id, "step": "weight"})
                await safe_answer(
                    message,
                    "⚖️ Jo'natmaning og'irligini kiriting.\n\n"
                    "Agar gabaritda o'lchangan og'irligi kattaroq bo'lsa, shuni kiriting. Masalan: 1.5",
                    reply_markup=reply_keyboard([], add_back=True),
                )
                return True
        except Exception as error:
            logger.exception("Calculator step failed")
            await safe_answer(message, f"⚠️ Kalkulyator xatoligi: {error}")
            return True

    if mode == "calculator" and state.get("step") == "weight":
        normalized = text.replace(",", ".")
        try:
            weight = float(normalized)
        except ValueError:
            await safe_answer(message, "Og'irlikni raqamda kiriting. Masalan: 1.5")
            return True
        if weight <= 0:
            await safe_answer(message, "Og'irlik 0 dan katta bo'lishi kerak.")
            return True

        try:
            sender_city_id = int(state["sender_city_id"])
            receiver_city_id = int(state["receiver_city_id"])
            service_id = int(state["service_id"])
            result = await calculate_emu_delivery(sender_city_id, receiver_city_id, weight, service_id)
            receiver_branches = await get_emu_branches(city_id=receiver_city_id)
            await safe_answer(
                message,
                format_calculator_result(result, service_id, receiver_branches),
                reply_markup=main_menu_keyboard(),
            )
            service_states.pop(message.chat.id, None)
        except Exception as error:
            logger.exception("Calculator failed")
            await safe_answer(message, f"Kalkulyator xatoligi: {error}")
        return True

    return False


async def clear_handler(message: Message, start_next_setup: bool = True) -> None:
    if not await ensure_user_access(message):
        return
    client_type = current_client_type(message.chat.id)
    async with excel_lock:
        reset_excel_file(client_type)
    sender_sessions.pop(message.chat.id, None)
    setup_states.pop(message.chat.id, None)
    await safe_answer(
        message,
        "Ro'yxat tozalandi. Yangi Excel fayl shablondan tayyorlanadi.",
        reply_markup=archive_menu_keyboard() if not start_next_setup else None,
    )
    if start_next_setup:
        await start_setup(message)


async def text_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    text = message.text or ""
    normalized_menu_text = MENU_ALIASES.get(text.strip(), text.strip())
    if normalized_menu_text in MENU_TEXTS:
        setup_states.pop(message.chat.id, None)
        if normalized_menu_text != MENU_BACK:
            service_states.pop(message.chat.id, None)
        await handle_menu_message(message)
        return

    if await handle_service_text(message):
        return

    if await handle_setup_message(message):
        return

    if await handle_menu_message(message):
        return

    if message.chat.id not in sender_sessions:
        if any(word in normalized_menu_text.casefold() for word in ["qancha", "narx", "kg", "кг", "so'm", "som"]):
            await answer_ai_question(message, text)
            return
        await show_collect_menu(message)
        return

    await enqueue_batch_item(BatchItem(kind="text", message=message, text=text))


async def photo_handler(message: Message, bot: Bot) -> None:
    if not await ensure_user_access(message):
        return
    if message.chat.id in setup_states:
        await safe_answer(message, "Hozir jo'natuvchi ma'lumotlarini matn ko'rinishida kiriting.")
        return

    if message.chat.id not in sender_sessions:
        await show_collect_menu(message)
        return

    photo = message.photo[-1]
    await enqueue_batch_item(
        BatchItem(
            kind="image",
            message=message,
            file_id=photo.file_id,
            mime_type="image/jpeg",
            bot=bot,
        )
    )


async def document_image_handler(message: Message, bot: Bot) -> None:
    if not await ensure_user_access(message):
        return
    if message.chat.id in setup_states:
        await safe_answer(message, "Hozir jo'natuvchi ma'lumotlarini matn ko'rinishida kiriting.")
        return

    if message.chat.id not in sender_sessions:
        await show_collect_menu(message)
        return

    document = message.document
    if document is None:
        await safe_answer(message, "Iltimos, rasm, Excel yoki mijoz ma'lumotlari yozilgan matn yuboring.")
        return

    file_name = document.file_name or ""
    mime_type = document.mime_type or ""
    if is_excel_document(file_name, mime_type):
        await enqueue_batch_item(
            BatchItem(
                kind="excel",
                message=message,
                file_id=document.file_id,
                mime_type=mime_type,
                file_name=file_name,
                bot=bot,
            )
        )
        return

    if file_name.casefold().endswith(".xls"):
        await safe_answer(message, "Excel faylni .xlsx formatida yuboring. Eski .xls format hozir qo'llab-quvvatlanmaydi.")
        return

    if not mime_type.startswith("image/"):
        await safe_answer(message, "Iltimos, rasm, .xlsx Excel yoki mijoz ma'lumotlari yozilgan matn yuboring.")
        return

    await enqueue_batch_item(
        BatchItem(
            kind="image",
            message=message,
            file_id=document.file_id,
            mime_type=mime_type or "image/jpeg",
            file_name=file_name,
            bot=bot,
        )
    )


async def unsupported_handler(message: Message) -> None:
    if not await ensure_user_access(message):
        return
    await safe_answer(message, "Matn, rasm yoki .xlsx Excel yuboring. Yordam uchun /help buyrug'ini bosing.")


async def setup_bot_commands(bot: Bot) -> None:
    await bot.set_my_commands(
        [
            BotCommand(command="start", description="Botni ishga tushirish"),
            BotCommand(command="setup", description="ФИЗ ЛИЦО jo'natuvchi sozlamalari"),
            BotCommand(command="ofislar", description="EMU ofislar ro'yxati"),
            BotCommand(command="kalkulyator", description="Yetkazib berish narxini hisoblash"),
            BotCommand(command="ai", description="EMU bo'yicha AI yordamchi"),
            BotCommand(command="help", description="Foydalanish bo'yicha yordam"),
            BotCommand(command="excel", description="Excel faylni yuborish"),
            BotCommand(command="shablon", description="Excel shablonni yuborish"),
            BotCommand(command="clear", description="Ro'yxatni tozalash"),
        ]
    )


async def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError(
            "BOT_TOKEN environment variable sozlanmagan. Railway > Variables bo'limiga BOT_TOKEN qo'shing."
        )
    if not OPENAI_API_KEY:
        raise RuntimeError(
            "OPENAI_API_KEY environment variable sozlanmagan. Railway > Variables bo'limiga OPENAI_API_KEY qo'shing."
        )

    ensure_excel_file(CLIENT_TYPE_PHYSICAL)
    ensure_excel_file(CLIENT_TYPE_LEGAL)
    if not emu_database_has_core_data() or not emu_database_is_fresh():
        logger.info("EMU local database is missing or stale, refreshing from API")
        await refresh_emu_database()

    bot = Bot(token=BOT_TOKEN)
    dispatcher = Dispatcher()
    await setup_bot_commands(bot)

    dispatcher.callback_query.register(access_callback_handler, F.data.startswith("access:"))
    dispatcher.callback_query.register(emu_callback_handler, F.data.startswith("emu:"))
    dispatcher.callback_query.register(setup_callback_handler, F.data.startswith("setup:"))
    dispatcher.message.register(start_handler, Command("start"))
    dispatcher.message.register(setup_handler, Command("setup"))
    dispatcher.message.register(offices_handler, Command("ofislar"))
    dispatcher.message.register(calculator_handler, Command("kalkulyator"))
    dispatcher.message.register(ai_handler, Command("ai"))
    dispatcher.message.register(help_handler, Command("help"))
    dispatcher.message.register(excel_handler, Command("excel"))
    dispatcher.message.register(template_handler, Command("shablon"))
    dispatcher.message.register(clear_handler, Command("clear"))
    dispatcher.message.register(access_contact_handler, F.contact)
    dispatcher.message.register(photo_handler, F.photo)
    dispatcher.message.register(document_image_handler, F.document)
    dispatcher.message.register(text_handler, F.text)
    dispatcher.message.register(unsupported_handler)

    logger.info("Bot started")
    await dispatcher.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot stopped")
